import json
import time
import requests
import pandas as pd
from datetime import datetime
import websocket
import threading
from urllib.parse import urlparse, parse_qs
import re
import os
import subprocess
import psutil
import webbrowser
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from io import BytesIO
import tempfile
import urllib.request
import platform
import shutil

class YouTubeStudioMonitor:
    def __init__(self, chrome_port=9222):
        self.chrome_port = chrome_port
        self.ws = None
        self.captured_request = None  # 캡처된 원본 요청 저장
        self.captured_analytics_request = None  # 캡처된 애널리틱스 요청 저장 (get_screen)
        self.captured_analytics_cards_request = None  # 캡처된 get_cards 요청 저장
        self.channel_id = None
        self.collected_data = []
        self.video_analytics_data = []  # 비디오 상세 정보 저장
        self.monitoring = False
        self.chrome_process = None
        self.pending_requests = {}  # 대기 중인 요청들 저장 (request_id -> request_data)
        self.collection_phase = "videos_list"  # "videos_list" → "video_analytics"
        
        # 다중 탭 애널리틱스 수집을 위한 설정
        self.analytics_tabs = [
            {
                'name': 'reach_viewers',
                'url_suffix': 'tab-reach_viewers/period-default',
                'api_endpoint': 'get_screen',
                'description': '노출수, CTR 데이터'
            },
            {
                'name': 'interest_viewers', 
                'url_suffix': 'tab-interest_viewers/period-default',
                'api_endpoint': 'get_cards',
                'description': '조회수, 시청시간, 구독자 증감 데이터'
            }
        ]
        self.current_tab_index = 0  # 현재 처리 중인 탭 인덱스
        self.collected_analytics_data = {}  # 탭별로 수집된 데이터 저장
        
        # 스레드 안전성을 위한 락
        import threading as thread_module
        self.tab_lock = thread_module.Lock()  # 탭 인덱스 변경 시 락
        
    def find_chrome_executable(self):
        """크롬 실행 파일 경로 찾기 (크로스 플랫폼)"""
        system = platform.system()
        
        if system == "Darwin":  # macOS
            possible_paths = [
                "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
                "/Applications/Google Chrome Canary.app/Contents/MacOS/Google Chrome Canary",
                "/Applications/Chromium.app/Contents/MacOS/Chromium",
                os.path.expanduser("~/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
            ]
        elif system == "Linux":
            possible_paths = [
                "/usr/bin/google-chrome",
                "/usr/bin/google-chrome-stable",
                "/usr/bin/chromium",
                "/usr/bin/chromium-browser",
                "/snap/bin/chromium",
                "/opt/google/chrome/chrome",
            ]
        elif system == "Windows":
            possible_paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe", 
                os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe"),
                r"C:\Users\%USERNAME%\AppData\Local\Google\Chrome\Application\chrome.exe"
            ]
        else:
            possible_paths = []
        
        # 경로에서 Chrome 찾기
        for path in possible_paths:
            expanded_path = os.path.expandvars(path)
            if os.path.exists(expanded_path):
                print(f"✅ Chrome 발견: {expanded_path}")
                return expanded_path
        
        # Windows에서는 레지스트리도 확인
        if system == "Windows":
            try:
                import winreg
                reg_path = r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\chrome.exe"
                with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, reg_path) as key:
                    chrome_path = winreg.QueryValue(key, "")
                    if os.path.exists(chrome_path):
                        print(f"✅ Chrome 발견 (레지스트리): {chrome_path}")
                        return chrome_path
            except Exception as e:
                print(f"레지스트리 검색 실패: {e}")
        
        # PATH에서 Chrome 찾기 (모든 플랫폼)
        chrome_names = []
        if system == "Darwin":
            # macOS에서는 PATH보다는 직접 경로를 우선시
            pass  
        elif system == "Linux":
            chrome_names = ["google-chrome", "google-chrome-stable", "chromium", "chromium-browser"]
        elif system == "Windows":
            chrome_names = ["chrome.exe", "chrome"]
        
        for name in chrome_names:
            chrome_path = shutil.which(name)
            if chrome_path:
                print(f"✅ Chrome 발견 (PATH): {chrome_path}")
                return chrome_path
        
        print(f"❌ Chrome을 찾을 수 없습니다 ({system})")
        return None
    
    def is_chrome_debug_running(self):
        """크롬이 디버그 모드로 실행 중인지 확인"""
        try:
            response = requests.get(f'http://localhost:{self.chrome_port}/json', timeout=2)
            return response.status_code == 200
        except:
            return False
    
    def kill_existing_chrome(self):
        """기존 크롬 프로세스 종료 (크로스 플랫폼)"""
        print("기존 크롬 프로세스를 종료하는 중...")
        system = platform.system()
        
        try:
            for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
                try:
                    proc_name = proc.info['name'].lower()
                    
                    # 플랫폼별 Chrome 프로세스 이름
                    chrome_names = []
                    if system == "Darwin":  # macOS
                        chrome_names = ['google chrome', 'chrome', 'chromium']
                    elif system == "Linux":
                        chrome_names = ['chrome', 'chromium', 'google-chrome', 'chromium-browser']
                    elif system == "Windows":
                        chrome_names = ['chrome.exe', 'chrome']
                    
                    # Chrome 프로세스인지 확인
                    is_chrome = any(name in proc_name for name in chrome_names)
                    
                    if is_chrome:
                        print(f"  Chrome 프로세스 종료: {proc.info['pid']} - {proc_name}")
                        proc.terminate()
                        
                except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                    pass
                    
        except Exception as e:
            print(f"프로세스 종료 중 오류: {e}")
        
        time.sleep(2)
    
    def get_program_profile_path(self):
        """프로그램 전용 Chrome 프로필 경로 반환 (크로스 플랫폼)"""
        system = platform.system()
        
        if system == "Darwin":  # macOS
            base_dir = os.path.expanduser("~/Library/Application Support")
        elif system == "Linux":
            base_dir = os.path.expanduser("~/.config")
        elif system == "Windows":
            base_dir = os.path.expanduser("~")
        else:
            base_dir = os.path.expanduser("~")
        
        profile_dir = os.path.join(base_dir, "YouTubeStudioMonitor", "chrome_profile")
        os.makedirs(profile_dir, exist_ok=True)
        print(f"Chrome 프로필 경로: {profile_dir}")
        return profile_dir
    
    def is_profile_logged_in(self, profile_path):
        """프로필이 Google 계정에 로그인되어 있는지 확인"""
        try:
            prefs_file = os.path.join(profile_path, "Default", "Preferences")
            if os.path.exists(prefs_file) and os.path.getsize(prefs_file) > 1000:
                return True
            return False
        except:
            return False
    
    def start_chrome_debug_mode(self):
        """크롬을 디버그 모드로 실행 (YouTube Studio로 바로 이동하지 않음)"""
        chrome_path = self.find_chrome_executable()
        if not chrome_path:
            print("크롬 실행 파일을 찾을 수 없습니다.")
            return False
        
        if not self.is_chrome_debug_running():
            self.kill_existing_chrome()
        
        user_data_dir = self.get_program_profile_path()
        is_logged_in = self.is_profile_logged_in(user_data_dir)
        
        if is_logged_in:
            print("기존 프로필을 사용합니다.")
        else:
            print("새 프로그램 전용 프로필을 생성합니다.")
        
        # YouTube Studio로 바로 이동하지 않고 Google 홈페이지로 시작
        chrome_args = [
            chrome_path,
            f"--remote-debugging-port={self.chrome_port}",
            f"--user-data-dir={user_data_dir}",
            f"--remote-allow-origins=http://localhost:{self.chrome_port}",
            "--no-first-run",
            "--no-default-browser-check",
            "--disable-web-security",
            "--disable-features=VizDisplayCompositor",
            "https://www.google.com"  # YouTube Studio 대신 Google 홈페이지로 시작
        ]
        
        try:
            print(f"크롬을 디버그 모드로 실행하는 중... (포트: {self.chrome_port})")
            self.chrome_process = subprocess.Popen(
                chrome_args,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=subprocess.CREATE_NEW_PROCESS_GROUP if os.name == 'nt' else 0
            )
            
            for i in range(10):
                if self.is_chrome_debug_running():
                    print("크롬이 성공적으로 시작되었습니다.")
                    return True
                time.sleep(1)
            
            print("크롬 시작 시간 초과")
            return False
            
        except Exception as e:
            print(f"크롬 실행 오류: {e}")
            return False

    def find_youtube_studio_tab(self):
        """YouTube Studio 탭 찾기"""
        try:
            response = requests.get(f'http://localhost:{self.chrome_port}/json')
            tabs = response.json()
            
            print(f"\n🔍 현재 열린 탭들을 확인 중... (총 {len(tabs)}개)")
            
            for i, tab in enumerate(tabs):
                url = tab.get('url', '')
                title = tab.get('title', '')
                print(f"  탭 {i+1}: {title[:50]}... - {url[:80]}...")
            
            studio_patterns = ['studio.youtube.com', 'youtube.com/studio', '/studio']
            
            for tab in tabs:
                url = tab.get('url', '').lower()
                for pattern in studio_patterns:
                    if pattern in url:
                        print(f"✅ YouTube Studio 탭 발견: {tab.get('title', 'Unknown')}")
                        return tab
            
            return None
            
        except Exception as e:
            print(f"❌ 탭 검색 중 오류: {e}")
            return None

    def extract_channel_id_from_url(self, url):
        """URL에서 채널 ID 추출"""
        try:
            # studio.youtube.com/channel/UCWA34FUr_rV6JFWId9RQg1A 패턴
            match = re.search(r'studio\.youtube\.com/channel/([A-Za-z0-9_-]+)', url)
            if match:
                return match.group(1)
            return None
        except:
            return None

    def navigate_to_videos_page(self):
        """비디오 목록 페이지로 이동"""
        try:
            if not self.channel_id:
                print("❌ 채널 ID가 없습니다.")
                return False
            
            # 비디오 업로드 페이지로 이동
            videos_url = f"https://studio.youtube.com/channel/{self.channel_id}/videos/upload"
            
            # 활성 탭에서 URL 변경
            response = requests.get(f'http://localhost:{self.chrome_port}/json')
            tabs = response.json()
            
            studio_tab = None
            for tab in tabs:
                if 'studio.youtube.com' in tab.get('url', ''):
                    studio_tab = tab
                    break
            
            if studio_tab:
                # JavaScript로 페이지 이동
                ws_url = studio_tab['webSocketDebuggerUrl']
                ws = websocket.create_connection(ws_url)
                
                ws.send(json.dumps({
                    "id": 1,
                    "method": "Runtime.evaluate",
                    "params": {
                        "expression": f"window.location.href = '{videos_url}'"
                    }
                }))
                
                ws.close()
                print(f"✅ 비디오 목록 페이지로 이동: {videos_url}")
                time.sleep(5)  # 페이지 로딩 대기
                return True
            
            return False
            
        except Exception as e:
            print(f"❌ 비디오 페이지 이동 오류: {e}")
            return False

    def navigate_to_video_analytics_page(self, video_id):
        """특정 비디오의 애널리틱스 페이지로 이동"""
        try:
            if not self.channel_id:
                print("❌ 채널 ID가 없습니다.")
                return False
            
            # 비디오 애널리틱스 페이지로 이동
            analytics_url = f"https://studio.youtube.com/video/{video_id}/analytics/tab-reach_viewers/period-default"
            
            # 활성 탭에서 URL 변경
            response = requests.get(f'http://localhost:{self.chrome_port}/json')
            tabs = response.json()
            
            studio_tab = None
            for tab in tabs:
                if 'studio.youtube.com' in tab.get('url', ''):
                    studio_tab = tab
                    break
            
            if studio_tab:
                # JavaScript로 페이지 이동
                ws_url = studio_tab['webSocketDebuggerUrl']
                ws = websocket.create_connection(ws_url)
                
                ws.send(json.dumps({
                    "id": 1,
                    "method": "Runtime.evaluate",
                    "params": {
                        "expression": f"window.location.href = '{analytics_url}'"
                    }
                }))
                
                ws.close()
                print(f"✅ 비디오 애널리틱스 페이지로 이동: {analytics_url}")
                time.sleep(8)  # 페이지 로딩 대기 (애널리틱스는 더 오래 걸림)
                return True
            
            return False
            
        except Exception as e:
            print(f"❌ 비디오 애널리틱스 페이지 이동 오류: {e}")
            return False

    def connect_to_chrome(self):
        """크롬 브라우저에 CDP로 연결"""
        try:
            if not self.is_chrome_debug_running():
                print("크롬이 디버그 모드로 실행되지 않았습니다. 자동으로 실행합니다...")
                
                if not self.start_chrome_debug_mode():
                    return False
                
                print("✅ Chrome이 시작되었습니다.")
                time.sleep(3)  # 짧은 로딩 대기
            
            print("📋 계정 설정 안내:")
            print("   1. 열린 Chrome 브라우저에서 원하는 Google 계정으로 로그인하세요")
            print("   2. YouTube Studio (https://studio.youtube.com) 페이지로 이동하세요")
            print("   3. 수집할 채널을 선택하고 비디오 목록 페이지까지 이동하세요")
            print("   4. 준비가 완료되면 이 터미널에서 엔터를 누르세요")
            
            # 사용자 입력 대기
            input("\n⏳ 로그인 및 채널 선택 완료 후 엔터를 누르세요...")
            
            print("\n🔍 YouTube Studio 탭을 찾는 중...")
            
            # YouTube Studio 탭 찾기 시도 (여러 번)
            target_tab = None
            for attempt in range(5):
                target_tab = self.find_youtube_studio_tab()
                if target_tab:
                    break
                print(f"   시도 {attempt + 1}/5: YouTube Studio 탭을 찾지 못했습니다. 다시 시도...")
                time.sleep(2)
            
            if not target_tab:
                print("❌ YouTube Studio 탭을 찾을 수 없습니다.")
                print("   💡 YouTube Studio (https://studio.youtube.com)로 이동했는지 확인해주세요.")
                return False
            
            # URL에서 채널 ID 추출
            studio_url = target_tab.get('url', '')
            self.channel_id = self.extract_channel_id_from_url(studio_url)
            
            if self.channel_id:
                print(f"✅ 채널 ID 추출됨: {self.channel_id}")
            else:
                print("⚠️ 채널 ID를 추출할 수 없습니다.")
                print("   💡 YouTube Studio의 채널 메인 페이지로 이동해주세요.")
            
            # WebSocket 연결
            ws_url = target_tab['webSocketDebuggerUrl']
            self.ws = websocket.create_connection(ws_url, timeout=10)
            
            # WebSocket 타임아웃 설정
            self.ws.settimeout(1.0)  # 1초 타임아웃
            
            # Network 도메인 활성화
            self.ws.send(json.dumps({
                "id": 1,
                "method": "Network.enable"
            }))
            
            # Runtime 도메인 활성화
            self.ws.send(json.dumps({
                "id": 2,
                "method": "Runtime.enable"
            }))
            
            print("✅ 크롬 브라우저에 성공적으로 연결되었습니다.")
            return True
            
        except Exception as e:
            print(f"❌ 크롬 연결 실패: {e}")
            return False
    
    def is_video_list_api(self, url):
        """비디오 목록 API 요청인지 확인"""
        return 'youtubei/v1/creator/list_creator_videos' in url
    
    def is_video_analytics_api(self, url):
        """비디오 애널리틱스 API 요청인지 확인 (get_screen 또는 get_cards)"""
        return 'youtubei/v1/yta_web/get_screen' in url or 'youtubei/v1/yta_web/get_cards' in url
    
    def is_get_screen_api(self, url):
        """get_screen API 요청인지 확인"""
        return 'youtubei/v1/yta_web/get_screen' in url
    
    def is_get_cards_api(self, url):
        """get_cards API 요청인지 확인"""
        return 'youtubei/v1/yta_web/get_cards' in url
    
    def fetch_cookies_for_url(self, url):
        """특정 URL에 대한 쿠키를 Chrome에서 가져오기"""
        try:
            if not self.ws:
                return None
            
            print(f"🍪 쿠키 정보 가져오는 중... (URL: {url})")
            
            # Network.getCookies 요청
            self.ws.send(json.dumps({
                "id": int(time.time()),
                "method": "Network.getCookies",
                "params": {"urls": [url]}
            }))
            
            # 응답 대기
            timeout = 5
            start_time = time.time()
            
            while time.time() - start_time < timeout:
                try:
                    message = self.ws.recv()
                    data = json.loads(message)
                    
                    if 'result' in data and 'cookies' in data['result']:
                        cookies = data['result']['cookies']
                        print(f"✅ 쿠키 획득: {len(cookies)}개")
                        
                        # Cookie 헤더 문자열 생성
                        cookie_pairs = []
                        for cookie in cookies:
                            name = cookie.get('name', '')
                            value = cookie.get('value', '')
                            if name and value:
                                cookie_pairs.append(f"{name}={value}")
                        
                        cookie_header = '; '.join(cookie_pairs)
                        print(f"🍪 Cookie 헤더 생성: {len(cookie_header)} bytes")
                        return cookie_header
                        
                except websocket.WebSocketTimeoutException:
                    continue
                except Exception as e:
                    print(f"쿠키 수신 오류: {e}")
                    break
            
            print("⏰ 쿠키 요청 시간 초과")
            return None
            
        except Exception as e:
            print(f"쿠키 요청 오류: {e}")
            return None

    def capture_request_data(self, request, request_id=None):
        """네트워크 요청 데이터 완전히 캡처"""
        try:
            captured = {
                'url': request['url'],
                'method': request['method'],
                'headers': request.get('headers', {}),
                'postData': request.get('postData', ''),
                'hasPostData': request.get('hasPostData', False),
                'request_id': request_id,
                'timestamp': datetime.now().isoformat()
            }
            
            print("🎯 비디오 목록 API 요청 완전 캡처!")
            print(f"   URL: {captured['url']}")
            print(f"   Method: {captured['method']}")
            print(f"   Headers: {len(captured['headers'])}개")
            print(f"   POST Data: {len(captured['postData'])} 바이트")
            print(f"   Has POST Data: {captured['hasPostData']}")
            
            # Cookie 헤더가 없다면 별도로 가져오기
            if 'Cookie' not in captured['headers'] and 'cookie' not in captured['headers']:
                print("⚠️ Cookie 헤더가 없습니다. 별도로 가져옵니다...")
                cookie_header = self.fetch_cookies_for_url(captured['url'])
                if cookie_header:
                    captured['headers']['Cookie'] = cookie_header
                    print(f"✅ Cookie 헤더 추가됨: {len(cookie_header)} bytes")
                else:
                    print("❌ 쿠키를 가져올 수 없습니다.")
            
            # POST 데이터가 있다고 표시되어 있는데 실제로는 없다면
            if captured['hasPostData'] and not captured['postData'] and request_id:
                print("⚠️ POST 데이터가 있다고 하는데 캡처되지 않음. 별도로 가져오기 시도...")
                captured['needs_post_data_fetch'] = True
            
            return captured
            
        except Exception as e:
            print(f"요청 캡처 오류: {e}")
            return None
    
    def fetch_post_data(self, request_id):
        """별도로 POST 데이터 가져오기"""
        try:
            if not self.ws or not request_id:
                return None
            
            print(f"📡 POST 데이터 별도 요청 중... (Request ID: {request_id})")
            
            # POST 데이터 요청
            self.ws.send(json.dumps({
                "id": int(time.time()),
                "method": "Network.getRequestPostData",
                "params": {"requestId": request_id}
            }))
            
            # 응답 대기 (간단한 동기 방식)
            timeout = 5
            start_time = time.time()
            
            while time.time() - start_time < timeout:
                try:
                    message = self.ws.recv()
                    data = json.loads(message)
                    
                    if 'result' in data and 'postData' in data['result']:
                        post_data = data['result']['postData']
                        print(f"✅ POST 데이터 획득: {len(post_data)} 바이트")
                        return post_data
                        
                except websocket.WebSocketTimeoutException:
                    continue
                except Exception as e:
                    print(f"POST 데이터 수신 오류: {e}")
                    break
            
            print("⏰ POST 데이터 요청 시간 초과")
            return None
            
        except Exception as e:
            print(f"POST 데이터 요청 오류: {e}")
            return None
    
    def process_network_request(self, message):
        """네트워크 요청 처리 - 성공한 응답을 받은 요청만 캡처"""
        try:
            method = message.get('method')
            params = message.get('params', {})
            
            # 요청이 보내질 때 - 일단 저장만 해둠
            if method == 'Network.requestWillBeSent':
                request = params['request']
                url = request['url']
                request_id = params.get('requestId')
                
                # 비디오 목록 API 요청 감지
                if self.is_video_list_api(url) and self.collection_phase == "videos_list":
                    print(f"🔍 비디오 목록 API 요청 감지! (대기 중...)")
                    print(f"   URL: {url}")
                    print(f"   Request ID: {request_id}")
                    
                    # 대기 중인 요청으로 저장 (아직 복제하지 않음)
                    self.pending_requests[request_id] = {
                        'request': request,
                        'request_id': request_id,
                        'timestamp': datetime.now().isoformat(),
                        'type': 'videos_list'
                    }
                    print(f"   ⏳ 응답 대기 중... (성공하면 복제할 예정)")
                
                # 비디오 애널리틱스 API 요청 감지
                elif self.is_video_analytics_api(url) and self.collection_phase == "video_analytics":
                    if self.is_get_screen_api(url):
                        api_type = "get_screen"
                        print(f"🎯 get_screen API 요청 감지! (대기 중...)")
                    elif self.is_get_cards_api(url):
                        api_type = "get_cards"
                        print(f"🎯 get_cards API 요청 감지! (대기 중...)")
                    else:
                        api_type = "unknown"
                        print(f"🎯 알 수 없는 애널리틱스 API 요청 감지! (대기 중...)")
                    
                    print(f"   URL: {url}")
                    print(f"   Request ID: {request_id}")
                    
                    # 대기 중인 요청으로 저장
                    self.pending_requests[request_id] = {
                        'request': request,
                        'request_id': request_id,
                        'timestamp': datetime.now().isoformat(),
                        'type': 'video_analytics',
                        'api_type': api_type
                    }
                    print(f"   ⏳ 응답 대기 중... (성공하면 복제할 예정)")
            
            # 응답을 받았을 때 - 성공한 경우에만 복제
            elif method == 'Network.responseReceived':
                request_id = params.get('requestId')
                response = params.get('response', {})
                status = response.get('status', 0)
                url = response.get('url', '')
                
                # 대기 중인 요청이고 성공한 경우
                if request_id in self.pending_requests:
                    pending_request = self.pending_requests[request_id]
                    request_type = pending_request.get('type')
                    
                    if status == 200:
                        if request_type == 'videos_list' and self.is_video_list_api(url):
                            print(f"✅ 비디오 목록 API 응답 성공! (상태코드: {status})")
                            self.handle_successful_videos_list_request(pending_request)
                            
                        elif request_type == 'video_analytics' and self.is_video_analytics_api(url):
                            api_type = pending_request.get('api_type', 'unknown')
                            if api_type == 'get_screen':
                                print(f"✅ get_screen API 응답 성공! (상태코드: {status})")
                                self.handle_successful_get_screen_request(pending_request)
                            elif api_type == 'get_cards':
                                print(f"✅ get_cards API 응답 성공! (상태코드: {status})")
                                self.handle_successful_get_cards_request(pending_request)
                            else:
                                print(f"✅ 알 수 없는 애널리틱스 API 응답 성공! (상태코드: {status})")
                                print(f"   URL: {url}")
                            self.handle_successful_analytics_request(pending_request)
                    else:
                        print(f"❌ API 응답 실패! (상태코드: {status})")
                        print(f"   실패한 요청은 복제하지 않습니다.")
                    
                    # 처리 완료된 요청은 제거
                    del self.pending_requests[request_id]
                    
        except Exception as e:
            print(f"네트워크 요청 처리 오류: {e}")
    
    def handle_successful_videos_list_request(self, pending_request):
        """성공한 비디오 목록 요청 처리"""
        try:
            print(f"   이제 성공한 비디오 목록 요청을 복제합니다.")
            
            # 성공한 요청 데이터 캡처
            self.captured_request = self.capture_request_data(
                pending_request['request'], 
                pending_request['request_id']
            )
            
            if self.captured_request:
                # POST 데이터가 별도로 필요한 경우
                request_id = pending_request['request_id']
                if self.captured_request.get('needs_post_data_fetch') and request_id:
                    additional_post_data = self.fetch_post_data(request_id)
                    if additional_post_data:
                        self.captured_request['postData'] = additional_post_data
                
                print("🚀 성공한 비디오 목록 요청을 복제하여 전송합니다!")
                # 별도 스레드에서 즉시 복제 요청 실행
                threading.Thread(
                    target=self.replay_captured_request, 
                    daemon=True
                ).start()
        except Exception as e:
            print(f"비디오 목록 요청 처리 오류: {e}")
    
    def handle_successful_analytics_request(self, pending_request):
        """성공한 애널리틱스 get_screen 요청 처리 - 다중 탭 시스템 사용으로 비활성화"""
        try:
            print(f"   ⚠️ 기존 단일 애널리틱스 처리 함수는 다중 탭 시스템으로 대체되었습니다.")
            print(f"   🔄 다중 탭 시스템이 이미 실행 중이므로 추가 처리를 건너뜁니다.")
            return
            
        except Exception as e:
            print(f"기존 애널리틱스 요청 처리 오류: {e}")
    
    def handle_successful_get_screen_request(self, pending_request):
        """성공한 get_screen API 요청 처리"""
        try:
            print(f"   이제 성공한 get_screen 요청을 캡처합니다.")
            
            # 성공한 요청 데이터 캡처
            captured_request = self.capture_request_data(
                pending_request['request'], 
                pending_request['request_id']
            )
            
            if captured_request:
                # POST 데이터가 별도로 필요한 경우
                request_id = pending_request['request_id']
                if captured_request.get('needs_post_data_fetch') and request_id:
                    additional_post_data = self.fetch_post_data(request_id)
                    if additional_post_data:
                        captured_request['postData'] = additional_post_data
                
                # 스레드 안전하게 탭 정보 가져오기
                with self.tab_lock:
                    # 인덱스 범위 체크 - 범위 초과 시 완료 처리
                    if self.current_tab_index >= len(self.analytics_tabs) or self.current_tab_index < 0:
                        print(f"🎉 모든 탭 수집 완료! (현재 인덱스: {self.current_tab_index}, 최대: {len(self.analytics_tabs) - 1})")
                        # 현재 비디오 ID 추출하고 완료 처리
                        current_video_id = self.extract_video_id_from_current_analytics()
                        if current_video_id:
                            self.finalize_video_analytics(current_video_id)
                        return
                    
                    # 현재 탭 데이터로 저장
                    current_tab = self.analytics_tabs[self.current_tab_index]
                    tab_name = current_tab['name']
                
                print(f"✅ {tab_name} 탭 데이터 캡처 완료! (탭 인덱스: {self.current_tab_index})")
                self.collected_analytics_data[tab_name] = {
                    'api_type': 'get_screen',
                    'captured_request': captured_request,
                    'tab_config': current_tab
                }
                
                # 현재 비디오 ID 추출 (URL에서)
                current_video_id = self.extract_video_id_from_current_analytics()
                
                # 다음 탭으로 진행
                print(f"🔄 다음 탭으로 진행합니다...")
                self.proceed_to_next_tab(current_video_id)
                
        except Exception as e:
            print(f"get_screen 요청 처리 오류: {e}")
            import traceback
            traceback.print_exc()
    
    def handle_successful_get_cards_request(self, pending_request):
        """성공한 get_cards API 요청 처리"""
        try:
            print(f"   이제 성공한 get_cards 요청을 캡처합니다.")
            
            # 성공한 요청 데이터 캡처
            captured_request = self.capture_request_data(
                pending_request['request'], 
                pending_request['request_id']
            )
            
            if captured_request:
                # POST 데이터가 별도로 필요한 경우
                request_id = pending_request['request_id']
                if captured_request.get('needs_post_data_fetch') and request_id:
                    additional_post_data = self.fetch_post_data(request_id)
                    if additional_post_data:
                        captured_request['postData'] = additional_post_data
                
                # 스레드 안전하게 탭 정보 가져오기
                with self.tab_lock:
                    # 인덱스 범위 체크 - 범위 초과 시 완료 처리
                    if self.current_tab_index >= len(self.analytics_tabs) or self.current_tab_index < 0:
                        print(f"🎉 모든 탭 수집 완료! (현재 인덱스: {self.current_tab_index}, 최대: {len(self.analytics_tabs) - 1})")
                        # 현재 비디오 ID 추출하고 완료 처리
                        current_video_id = self.extract_video_id_from_current_analytics()
                        if current_video_id:
                            self.finalize_video_analytics(current_video_id)
                        return
                    
                    # 현재 탭 데이터로 저장
                    current_tab = self.analytics_tabs[self.current_tab_index]
                    tab_name = current_tab['name']
                
                print(f"✅ {tab_name} 탭 데이터 캡처 완료! (탭 인덱스: {self.current_tab_index})")
                self.collected_analytics_data[tab_name] = {
                    'api_type': 'get_cards',
                    'captured_request': captured_request,
                    'tab_config': current_tab
                }
                
                # 현재 비디오 ID 추출
                current_video_id = self.extract_video_id_from_current_analytics()
                
                # 다음 탭으로 진행
                print(f"🔄 다음 탭으로 진행합니다...")
                self.proceed_to_next_tab(current_video_id)
                
        except Exception as e:
            print(f"get_cards 요청 처리 오류: {e}")
            import traceback
            traceback.print_exc()
    
    def extract_video_id_from_current_analytics(self):
        """현재 브라우저의 애널리틱스 URL에서 비디오 ID 추출"""
        try:
            response = requests.get(f'http://localhost:{self.chrome_port}/json')
            tabs = response.json()
            
            for tab in tabs:
                url = tab.get('url', '')
                if 'studio.youtube.com/video/' in url and '/analytics/' in url:
                    # URL에서 비디오 ID 추출: .../video/VIDEO_ID/analytics/...
                    match = re.search(r'/video/([A-Za-z0-9_-]+)/analytics/', url)
                    if match:
                        video_id = match.group(1)
                        print(f"🔍 현재 비디오 ID: {video_id}")
                        return video_id
            
            print("⚠️ 현재 비디오 ID를 추출할 수 없습니다.")
            return None
            
        except Exception as e:
            print(f"❌ 비디오 ID 추출 오류: {e}")
            return None

    def generate_curl_command(self, url, method, headers, post_data):
        """디버깅용 CURL 명령어 생성"""
        try:
            curl_parts = [f"curl --location '{url}'"]
            
            # 헤더 추가
            for key, value in headers.items():
                # 특수 문자 이스케이프
                escaped_value = str(value).replace("'", "'\"'\"'").replace('"', '\\"')
                curl_parts.append(f"--header '{key}: {escaped_value}'")
            
            # POST 데이터 추가
            if method.upper() == 'POST' and post_data:
                # JSON 데이터 이스케이프
                escaped_data = post_data.replace("'", "'\"'\"'").replace('"', '\\"')
                curl_parts.append(f"--data '{escaped_data}'")
            
            return " \\\n".join(curl_parts)
        except Exception as e:
            return f"CURL 생성 오류: {e}"
    
    def save_debug_info(self, original_headers, replicated_headers, curl_command):
        """디버깅 정보를 파일로 저장"""
        debug_info = {
            "timestamp": datetime.now().isoformat(),
            "original_headers": original_headers,
            "replicated_headers": replicated_headers,
            "curl_command": curl_command,
            "comparison": {
                "original_header_count": len(original_headers),
                "replicated_header_count": len(replicated_headers),
                "missing_headers": [],
                "different_headers": []
            }
        }
        
        # 헤더 비교
        for key, value in original_headers.items():
            if key not in replicated_headers:
                debug_info["comparison"]["missing_headers"].append(key)
            elif replicated_headers.get(key) != value:
                debug_info["comparison"]["different_headers"].append({
                    "header": key,
                    "original": value,
                    "replicated": replicated_headers.get(key)
                })
        
        # 디버그 파일 저장
        debug_filename = f"debug_request_{int(time.time())}.json"
        with open(debug_filename, 'w', encoding='utf-8') as f:
            json.dump(debug_info, f, ensure_ascii=False, indent=2)
        
        # CURL 명령어 파일로도 저장
        curl_filename = f"debug_curl_{int(time.time())}.sh"
        with open(curl_filename, 'w', encoding='utf-8') as f:
            f.write("#!/bin/bash\n")
            f.write("# 복제된 YouTube Studio API 요청\n")
            f.write("# 이 명령어로 직접 테스트할 수 있습니다\n\n")
            f.write(curl_command)
        
        return debug_filename, curl_filename

    def replay_captured_request(self):
        """캡처된 요청을 그대로 복제해서 다시 보내기 + 페이지네이션 처리"""
        try:
            if not self.captured_request:
                print("❌ 캡처된 요청이 없습니다.")
                return None
            
            print("\n🚀 캡처된 요청을 복제하여 전송 중...")
            
            # 캡처된 정보 그대로 사용
            url = self.captured_request['url']
            method = self.captured_request['method']
            original_headers = self.captured_request['headers'].copy()
            original_post_data = self.captured_request['postData']
            
            print(f"📋 원본 요청 정보:")
            print(f"   URL: {url}")
            print(f"   Method: {method}")
            print(f"   Headers: {len(original_headers)}개")
            print(f"   POST Data: {len(original_post_data)} bytes")
            
            # 모든 비디오를 저장할 리스트
            all_videos = []
            page_count = 0
            next_page_token = None
            
            while True:
                page_count += 1
                print(f"\n📄 페이지 {page_count} 수집 중...")
                
                # requests 세션 생성
                session = requests.Session()
                headers = original_headers.copy()
                post_data = original_post_data
                
                # 두 번째 페이지부터는 pageToken 추가
                if next_page_token:
                    try:
                        payload = json.loads(original_post_data)
                        payload['pageToken'] = next_page_token
                        post_data = json.dumps(payload, separators=(',', ':'))
                        print(f"   🔄 페이지 토큰 추가: {next_page_token[:50]}...")
                    except json.JSONDecodeError:
                        print(f"   ❌ 페이로드 파싱 실패")
                        break
                
                # 쿠키 설정
                if 'Cookie' in headers:
                    cookie_str = headers['Cookie']
                    cookies = {}
                    for item in cookie_str.split(';'):
                        if '=' in item:
                            key, value = item.strip().split('=', 1)
                            cookies[key] = value
                    session.cookies.update(cookies)
                    print(f"   🍪 쿠키 설정: {len(cookies)}개")
                
                # Content-Length 제거
                headers.pop('Content-Length', None)
                
                # 요청 전송
                print(f"   📡 API 요청 전송 중...")
                if method.upper() == 'POST':
                    if post_data:
                        try:
                            json_data = json.loads(post_data)
                            response = session.post(url, headers=headers, json=json_data, timeout=30)
                        except json.JSONDecodeError:
                            response = session.post(url, headers=headers, data=post_data, timeout=30)
                    else:
                        response = session.post(url, headers=headers, timeout=30)
                else:
                    response = session.get(url, headers=headers, timeout=30)
                
                print(f"   📨 응답 수신: {response.status_code}")
                
                if response.status_code == 200:
                    try:
                        api_response = response.json()
                        
                        # 이 페이지의 비디오들 추가
                        page_videos = []
                        if 'videos' in api_response:
                            page_videos = api_response['videos']
                        elif 'video' in api_response:
                            page_videos = api_response['video']
                        elif 'items' in api_response:
                            page_videos = api_response['items']
                        
                        print(f"   ✅ 페이지 {page_count}: {len(page_videos)}개 비디오 수집")
                        all_videos.extend(page_videos)
                        
                        # 다음 페이지 토큰 확인
                        next_page_token = api_response.get('nextPageToken')
                        if next_page_token:
                            print(f"   🔄 다음 페이지 토큰 발견: {next_page_token[:50]}...")
                            print(f"   ➡️ 다음 페이지로 진행합니다...")
                        else:
                            print(f"   🏁 마지막 페이지입니다! (nextPageToken 없음)")
                            break
                            
                    except json.JSONDecodeError:
                        print(f"   ❌ JSON 파싱 실패: {response.text[:200]}...")
                        break
                else:
                    print(f"   ❌ 요청 실패: {response.status_code}")
                    print(f"   응답: {response.text[:300]}...")
                    break
                
                # 페이지 간 짧은 대기 (API 제한 고려)
                if next_page_token:
                    time.sleep(1)
            
            # 전체 결과 처리
            if all_videos:
                print(f"\n🎉 전체 페이지네이션 수집 완료!")
                print(f"   📄 총 페이지: {page_count}개")
                print(f"   📹 총 비디오: {len(all_videos)}개")
                
                # 마지막 응답 구조를 베이스로 전체 결과 생성
                final_response = {
                    'videos': all_videos,
                    'total_videos': len(all_videos),
                    'pages_collected': page_count,
                    'collection_method': 'pagination'
                }
                
                self.parse_and_save_video_data(final_response)
                return final_response
            else:
                print(f"❌ 수집된 비디오가 없습니다.")
                return None
                
        except Exception as e:
            print(f"❌ 요청 복제 오류: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def parse_and_save_video_data(self, api_response):
        """API 응답에서 비디오 데이터 파싱 및 저장 (공개 영상만)"""
        try:
            if not api_response:
                print("❌ API 응답이 없습니다.")
                return
            
            # 응답 구조 확인
            print("📋 API 응답 구조 분석:")
            print(f"   최상위 키들: {list(api_response.keys())}")
            
            # videos 또는 video 키 찾기
            videos = []
            if 'videos' in api_response:
                videos = api_response['videos']
            elif 'video' in api_response:
                videos = api_response['video']
            elif 'items' in api_response:
                videos = api_response['items']
            else:
                # 다른 가능한 키들도 확인
                for key, value in api_response.items():
                    if isinstance(value, list) and len(value) > 0:
                        if isinstance(value[0], dict) and 'videoId' in value[0]:
                            videos = value
                            print(f"   비디오 목록을 '{key}' 키에서 발견!")
                            break
            
            if not videos:
                print("⚠️ 응답에서 비디오 목록을 찾을 수 없습니다.")
                # 전체 응답을 파일로 저장해서 구조 확인
                debug_filename = f"debug_response_{int(time.time())}.json"
                with open(debug_filename, 'w', encoding='utf-8') as f:
                    json.dump(api_response, f, ensure_ascii=False, indent=2)
                print(f"   디버깅을 위해 전체 응답을 {debug_filename}에 저장했습니다.")
                return
            
            print(f"✅ 총 {len(videos)}개 비디오를 발견했습니다.")
            
            processed_videos = []
            skipped_videos = []
            
            for i, video in enumerate(videos):
                try:
                    # 공개 상태 확인
                    privacy = video.get('privacy', '')
                    status = video.get('status', '')
                    
                    # 공개 영상만 처리 (YouTube API의 실제 값 사용)
                    if privacy != 'VIDEO_PRIVACY_PUBLIC':
                        video_title = video.get('title', '제목없음')[:30]
                        skipped_videos.append({
                            'title': video_title,
                            'privacy': privacy,
                            'videoId': video.get('videoId', 'Unknown')
                        })
                        print(f"  ⚠️ 제외: {video_title}... (상태: {privacy})")
                        continue
                    
                    # 동적으로 필드 추출 - 존재하는 필드만 가져오기
                    video_data = {
                        'collected_at': datetime.now().isoformat(),
                        'raw_data': video  # 원본 데이터도 보존
                    }
                    
                    # 기본 필드들 추출 (있는 것만)
                    basic_fields = [
                        'videoId', 'title', 'description', 'privacy', 
                        'lengthSeconds', 'timeCreatedSeconds', 'timePublishedSeconds',
                        'status', 'watchUrl', 'shareUrl'
                    ]
                    
                    for field in basic_fields:
                        if field in video:
                            video_data[field] = video[field]
                    
                    # 지표 데이터 추출
                    if 'publicMetrics' in video:
                        metrics = video['publicMetrics']
                        for metric_key, metric_value in metrics.items():
                            video_data[f'public_{metric_key}'] = metric_value
                    
                    if 'privateMetrics' in video:
                        metrics = video['privateMetrics']
                        for metric_key, metric_value in metrics.items():
                            video_data[f'private_{metric_key}'] = metric_value
                    
                    # 썸네일 정보
                    if 'thumbnailDetails' in video and 'thumbnails' in video['thumbnailDetails']:
                        thumbnails = video['thumbnailDetails']['thumbnails']
                        video_data['thumbnail_urls'] = [thumb.get('url') for thumb in thumbnails if 'url' in thumb]
                    
                    processed_videos.append(video_data)
                    
                    # 간단한 정보 출력
                    title = video_data.get('title', '제목없음')[:50]
                    view_count = video_data.get('public_viewCount', video_data.get('viewCount', '?'))
                    privacy_status = video_data.get('privacy', 'Unknown')
                    print(f"  📹 {len(processed_videos)}. {title}... (조회수: {view_count}, 상태: {privacy_status})")
                    
                except Exception as e:
                    print(f"⚠️ 비디오 {i+1} 처리 중 오류: {e}")
                    continue
            
            # 필터링 결과 요약
            print(f"\n📊 비디오 필터링 결과:")
            print(f"   ✅ 공개 영상 (VIDEO_PRIVACY_PUBLIC): {len(processed_videos)}개")
            print(f"   ⚠️ 제외된 영상: {len(skipped_videos)}개")
            
            # 제외된 영상을 상태별로 그룹화
            if skipped_videos:
                privacy_counts = {}
                for skipped in skipped_videos:
                    privacy_status = skipped['privacy']
                    if privacy_status not in privacy_counts:
                        privacy_counts[privacy_status] = 0
                    privacy_counts[privacy_status] += 1
                
                print(f"   제외된 영상 상태별 분류:")
                privacy_labels = {
                    'VIDEO_PRIVACY_PRIVATE': '비공개',
                    'VIDEO_PRIVACY_UNLISTED': '제한공개 (링크를 아는 사람만)',
                    'VIDEO_PRIVACY_DRAFT': '임시보관함',
                    'VIDEO_PRIVACY_SCHEDULED': '예약 게시'
                }
                
                for privacy_status, count in privacy_counts.items():
                    label = privacy_labels.get(privacy_status, privacy_status)
                    print(f"     • {label}: {count}개")
                
                print(f"   제외된 영상 목록 (처음 5개):")
                for skipped in skipped_videos[:5]:  # 처음 5개만 표시
                    privacy_status = skipped['privacy']
                    label = privacy_labels.get(privacy_status, privacy_status)
                    print(f"     • {skipped['title']}... ({label})")
                if len(skipped_videos) > 5:
                    print(f"     ... 및 {len(skipped_videos) - 5}개 더")
            
            if not processed_videos:
                print("❌ 공개 상태인 비디오가 없습니다.")
                return []
            
            # JSON 및 엑셀 파일로 저장
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # JSON 저장
            json_filename = f"youtube_videos_list_{timestamp}.json"
            save_data = {
                'channel_id': self.channel_id,
                'total_videos': len(processed_videos),
                'total_skipped': len(skipped_videos),
                'collected_at': datetime.now().isoformat(),
                'original_response_keys': list(api_response.keys()),
                'videos': processed_videos,
                'skipped_videos': skipped_videos
            }
            
            with open(json_filename, 'w', encoding='utf-8') as f:
                json.dump(save_data, f, ensure_ascii=False, indent=2)
            
            # 엑셀 저장 제거 (최종 애널리틱스 파일에서 통합 처리)
            
            print(f"🎉 공개 비디오 목록이 저장되었습니다!")
            print(f"   📄 JSON: {json_filename}")
            print(f"📊 총 {len(processed_videos)}개 공개 비디오 정보 저장됨")
            print(f"📌 최종 엑셀 파일은 애널리틱스 수집 완료 후 생성됩니다.")
            
            # 메모리에도 저장
            self.collected_data = processed_videos
            
            # 비디오 목록 수집 완료! 이제 애널리틱스 수집 단계로 이동
            print(f"✅ 1단계 완료: 공개 비디오 목록 수집 완료!")
            print(f"🎯 2단계 시작: 각 비디오의 상세 정보(애널리틱스) 수집을 시작합니다.")
            self.collection_phase = "video_analytics"
            
            # 첫 번째 비디오의 다중 탭 애널리틱스 수집 시작
            if processed_videos and len(processed_videos) > 0:
                first_video_id = processed_videos[0].get('videoId')
                if first_video_id:
                    print(f"📊 첫 번째 비디오 다중 탭 애널리틱스 수집 시작: {first_video_id}")
                    print(f"🎯 수집할 탭들: {[tab['name'] for tab in self.analytics_tabs]}")
                    self.start_multi_tab_analytics_collection(first_video_id)
                else:
                    print("❌ 첫 번째 비디오 ID를 찾을 수 없습니다.")
            else:
                print("❌ 수집된 공개 비디오가 없습니다.")
            
            return processed_videos
            
        except Exception as e:
            print(f"❌ 비디오 데이터 파싱 오류: {e}")
            # 오류 발생시에도 원본 응답 저장
            if api_response:
                error_filename = f"error_response_{int(time.time())}.json"
                with open(error_filename, 'w', encoding='utf-8') as f:
                    json.dump(api_response, f, ensure_ascii=False, indent=2)
                print(f"   오류 분석을 위해 응답을 {error_filename}에 저장했습니다.")
            return None
    
    def listen_for_messages(self):
        """WebSocket 메시지 수신 대기"""
        print("👂 네트워크 요청 모니터링 시작...")
        while self.monitoring:
            try:
                message = self.ws.recv()
                data = json.loads(message)
                
                if 'method' in data:
                    self.process_network_request(data)
                    
            except websocket.WebSocketConnectionClosedException:
                print("WebSocket 연결이 끊어졌습니다.")
                break
            except websocket.WebSocketTimeoutException:
                # 타임아웃은 정상적인 상황 - 조용히 계속 진행
                continue
            except Exception as e:
                if self.monitoring:  # 모니터링 중일 때만 오류 출력
                    print(f"메시지 수신 오류: {e}")
    
    def start_monitoring(self, duration=300):
        """모니터링 시작"""
        if not self.connect_to_chrome():
            return
        
        print(f"\n🔍 YouTube Studio 다중 탭 애널리틱스 수집 시작 ({duration}초 동안)")
        print("📋 수행할 작업:")
        print("   🔸 1단계: 비디오 목록 페이지에서 list_creator_videos API 감지")
        print("   🔸 2단계: 다중 탭 애널리틱스 수집")
        print("     • tab-reach_viewers → get_screen API 감지 (노출수, CTR)")
        print("     • tab-interest_viewers → get_cards API 감지 (조회수, 시청시간, 구독자)")
        print("   🔸 3단계: 모든 비디오의 상세 정보 자동 수집")
        print("   🔸 4단계: 완전한 데이터를 JSON 파일로 저장")
        
        if self.channel_id:
            print("\n📌 다음 단계:")
            print("   1. YouTube Studio에서 '콘텐츠' 메뉴를 클릭하여 비디오 목록으로 이동하세요")
            print("   2. 페이지가 로딩되면 자동으로 API 요청이 감지됩니다")
            print("   3. 비디오 목록 수집 후 자동으로 각 비디오의 애널리틱스를 수집합니다")
        else:
            print("\n⚠️ 채널 ID가 감지되지 않았습니다.")
            print("   YouTube Studio의 채널 메인 페이지로 이동한 후 '콘텐츠' 메뉴를 클릭해주세요!")
        
        self.monitoring = True
        
        # 메시지 수신 스레드 시작
        listener_thread = threading.Thread(target=self.listen_for_messages)
        listener_thread.daemon = True
        listener_thread.start()
        
        # 지정된 시간 동안 대기
        print(f"\n⏰ {duration}초 동안 API 요청을 감지합니다...")
        print(f"📌 1단계 완료 후 자동으로 2단계로 진행됩니다.")
        time.sleep(duration)
        
        # 모니터링 중단
        self.monitoring = False
        print("\n⏹️ 모니터링을 중단합니다.")
        
        if self.captured_request and self.video_analytics_data:
            print("🎉 모든 데이터 수집이 성공적으로 완료되었습니다!")
            print(f"   ✅ 비디오 목록: {len(self.collected_data)}개")
            print(f"   ✅ 애널리틱스: {len(self.video_analytics_data)}개")
        elif self.captured_request:
            print("✅ 1단계(비디오 목록)는 완료되었지만 2단계가 완료되지 않았습니다.")
            print("   수동으로 비디오 애널리틱스 페이지를 방문해보세요.")
        else:
            print("❌ API 요청이 감지되지 않았습니다.")
            print("   수동으로 YouTube Studio에서 비디오 목록 페이지를 새로고침해보세요.")
        
        # 연결 종료
        if self.ws:
            self.ws.close()

    def cleanup(self):
        """리소스 정리"""
        if self.ws:
            self.ws.close()
        
        print("\n✅ 모니터링이 완료되었습니다.")
        print("💡 Chrome 브라우저는 로그인 상태 유지를 위해 계속 실행됩니다.")

    def collect_all_video_analytics(self):
        """수집된 모든 비디오의 애널리틱스 데이터 수집"""
        try:
            if not self.captured_analytics_request:
                print("❌ 캡처된 get_screen 애널리틱스 요청이 없습니다.")
                return
            
            if not self.collected_data or len(self.collected_data) == 0:
                print("❌ 수집된 비디오 목록이 없습니다.")
                return
            
            print(f"\n🎯 모든 비디오의 애널리틱스 데이터 수집 시작!")
            print(f"📊 총 {len(self.collected_data)}개 비디오의 상세 정보를 수집합니다.")
            
            # 원본 요청 데이터
            url = self.captured_analytics_request['url']
            method = self.captured_analytics_request['method']
            headers = self.captured_analytics_request['headers'].copy()
            original_post_data = self.captured_analytics_request['postData']
            
            if not original_post_data:
                print("❌ get_screen 요청에 POST 데이터가 없습니다.")
                return
            
            # JSON 페이로드 파싱
            try:
                original_payload = json.loads(original_post_data)
            except json.JSONDecodeError:
                print("❌ get_screen 요청 페이로드를 JSON으로 파싱할 수 없습니다.")
                return
            
            print(f"🔧 원본 페이로드 구조 확인...")
            print(f"   페이로드 크기: {len(original_post_data)} bytes")
            
            successful_analytics = []
            failed_analytics = []
            
            # 각 비디오에 대해 애널리틱스 요청
            for i, video in enumerate(self.collected_data):
                video_id = video.get('videoId')
                video_title = video.get('title', '제목없음')[:30]
                
                if not video_id:
                    print(f"⚠️ {i+1}번째 비디오에 videoId가 없습니다.")
                    failed_analytics.append({'video': video, 'reason': 'No videoId'})
                    continue
                
                print(f"\n📹 [{i+1}/{len(self.collected_data)}] {video_title}... (ID: {video_id})")
                
                # 페이로드에서 비디오 ID 교체
                modified_payload = self.replace_video_id_in_payload(original_payload, video_id)
                
                if not modified_payload:
                    print(f"❌ 페이로드에서 비디오 ID를 교체할 수 없습니다.")
                    failed_analytics.append({'video': video, 'reason': 'Payload modification failed'})
                    continue
                
                # 수정된 페이로드로 요청 전송
                analytics_data = self.send_analytics_request(url, method, headers, modified_payload, video_id)
                
                if analytics_data:
                    # 성공적으로 수집된 데이터 저장
                    video_analytics = {
                        'video_id': video_id,
                        'video_title': video.get('title'),
                        'collected_at': datetime.now().isoformat(),
                        'analytics_data': analytics_data,
                        'basic_video_info': video
                    }
                    successful_analytics.append(video_analytics)
                    self.video_analytics_data.append(video_analytics)
                    print(f"✅ 애널리틱스 데이터 수집 성공!")
                else:
                    failed_analytics.append({'video': video, 'reason': 'Request failed'})
                    print(f"❌ 애널리틱스 데이터 수집 실패")
                
                # 요청 간 간격 (YouTube API 제한 고려)
                if i < len(self.collected_data) - 1:  # 마지막이 아니면
                    print(f"⏰ 다음 요청까지 3초 대기...")
                    time.sleep(3)
            
            # 결과 요약
            print(f"\n🎉 애널리틱스 데이터 수집 완료!")
            print(f"   ✅ 성공: {len(successful_analytics)}개")
            print(f"   ❌ 실패: {len(failed_analytics)}개")
            
            # 수집된 데이터 저장
            if successful_analytics:
                self.save_analytics_data(successful_analytics)
            
            # 모든 데이터 수집 완료 - 모니터링 중단
            print(f"✅ 모든 데이터 수집 완료! 모니터링을 중단합니다.")
            self.monitoring = False
            
        except Exception as e:
            print(f"❌ 애널리틱스 수집 오류: {e}")
            import traceback
            traceback.print_exc()
    
    def replace_video_id_in_payload(self, original_payload, new_video_id):
        """페이로드에서 비디오 ID를 새로운 ID로 교체"""
        try:
            # 페이로드 깊은 복사
            modified_payload = json.loads(json.dumps(original_payload))
            
            # 비디오 ID가 들어갈 수 있는 위치들 확인 및 교체
            replaced_count = 0
            
            # screenConfig.entity.videoId
            if 'screenConfig' in modified_payload and 'entity' in modified_payload['screenConfig']:
                if 'videoId' in modified_payload['screenConfig']['entity']:
                    modified_payload['screenConfig']['entity']['videoId'] = new_video_id
                    replaced_count += 1
            
            # screenConfig.timePeriod.entity.videoId
            if ('screenConfig' in modified_payload and 
                'timePeriod' in modified_payload['screenConfig'] and 
                'entity' in modified_payload['screenConfig']['timePeriod']):
                if 'videoId' in modified_payload['screenConfig']['timePeriod']['entity']:
                    modified_payload['screenConfig']['timePeriod']['entity']['videoId'] = new_video_id
                    replaced_count += 1
            
            # 재귀적으로 모든 videoId 찾기 및 교체
            def replace_video_ids_recursive(obj, target_id):
                nonlocal replaced_count
                if isinstance(obj, dict):
                    for key, value in obj.items():
                        if key == 'videoId' and isinstance(value, str):
                            obj[key] = target_id
                            replaced_count += 1
                        elif isinstance(value, (dict, list)):
                            replace_video_ids_recursive(value, target_id)
                elif isinstance(obj, list):
                    for item in obj:
                        replace_video_ids_recursive(item, target_id)
            
            replace_video_ids_recursive(modified_payload, new_video_id)
            
            if replaced_count > 0:
                print(f"   🔄 페이로드에서 {replaced_count}개 videoId를 교체했습니다.")
                return modified_payload
            else:
                print(f"   ⚠️ 페이로드에서 videoId를 찾지 못했습니다.")
                return None
                
        except Exception as e:
            print(f"   ❌ 페이로드 수정 오류: {e}")
            return None
    
    def send_analytics_request(self, url, method, headers, payload, video_id):
        """수정된 페이로드로 애널리틱스 요청 전송"""
        try:
            # requests 세션 생성
            session = requests.Session()
            request_headers = headers.copy()
            
            # 쿠키 설정
            if 'Cookie' in request_headers:
                cookie_str = request_headers['Cookie']
                cookies = {}
                for item in cookie_str.split(';'):
                    if '=' in item:
                        key, value = item.strip().split('=', 1)
                        cookies[key] = value
                session.cookies.update(cookies)
            
            # Content-Length 제거 (requests가 자동 계산)
            request_headers.pop('Content-Length', None)
            
            # JSON 요청 전송
            json_payload = json.dumps(payload, separators=(',', ':'))
            print(f"   📡 애널리틱스 요청 전송 중... (페이로드: {len(json_payload)} bytes)")
            
            response = session.post(
                url, 
                headers=request_headers, 
                json=payload, 
                timeout=30
            )
            
            if response.status_code == 200:
                try:
                    analytics_data = response.json()
                    print(f"   ✅ 응답 성공 (크기: {len(response.text)} bytes)")
                    return analytics_data
                except json.JSONDecodeError:
                    print(f"   ❌ 응답이 JSON 형식이 아닙니다.")
                    return None
            else:
                print(f"   ❌ 요청 실패: {response.status_code}")
                return None
                
        except Exception as e:
            print(f"   ❌ 애널리틱스 요청 오류: {e}")
            return None
    
    def save_analytics_data(self, analytics_data):
        """애널리틱스 데이터를 JSON과 간단한 엑셀 파일로 저장"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # JSON 저장
            json_filename = f"youtube_analytics_data_{timestamp}.json"
            save_data = {
                'channel_id': self.channel_id,
                'total_videos_analyzed': len(analytics_data),
                'collected_at': datetime.now().isoformat(),
                'analytics_data': analytics_data
            }
            
            with open(json_filename, 'w', encoding='utf-8') as f:
                json.dump(save_data, f, ensure_ascii=False, indent=2)
            
            # 간단한 엑셀 저장 (하나의 시트만)
            excel_filename = f"youtube_analytics_data_{timestamp}.xlsx"
            self.save_simple_analytics_excel(analytics_data, excel_filename)
            
            print(f"📊 애널리틱스 데이터가 저장되었습니다!")
            print(f"   📄 JSON: {json_filename}")
            print(f"   📊 Excel: {excel_filename}")
            
            # 간단한 요약 출력
            print(f"\n📈 수집된 데이터 요약:")
            for item in analytics_data[:3]:  # 처음 3개만 미리보기
                video_title = item.get('video_title', '제목없음')[:30]
                video_id = item.get('video_id', 'Unknown')
                print(f"   📹 {video_title}... (ID: {video_id})")
            
            if len(analytics_data) > 3:
                print(f"   ... 및 {len(analytics_data) - 3}개 더")
                
        except Exception as e:
            print(f"❌ 애널리틱스 데이터 저장 오류: {e}")
    
    def save_simple_analytics_excel(self, analytics_data, filename):
        """간단한 애널리틱스 엑셀 파일 생성 (단일 시트)"""
        try:
            if not analytics_data:
                print("❌ 저장할 애널리틱스 데이터가 없습니다.")
                return

            # 워크북 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'YouTube Analytics'
            
            # 헤더 설정 (썸네일을 첫 번째 열로, 단순화)
            headers = [
                '썸네일', '비디오 ID', '제목', '공개상태', '길이', '게시일',
                '조회수', '노출수', '클릭률', 
                '시청시간 (분)', '평균 시청시간 (초)', 
                '평균 조회율 (%)'
            ]
            
            # 헤더 작성
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num, value=header)
                cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
                cell.font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            
            # 데이터 처리
            for row_idx, item in enumerate(analytics_data, 2):
                video_info = item.get('basic_video_info', {})
                video_id = item.get('video_id', 'N/A')
                video_title = item.get('video_title', 'N/A')
                
                # 다중 탭 API 응답 데이터에서 메트릭 추출
                tabs_api_data = item.get('analytics_data', {})
                
                impressions = 0
                views = 0
                click_rate = 0
                watch_time_minutes = 0
                average_view_duration_seconds = 0
                average_percentage_watched = 0
                
                # 각 탭의 API 응답에서 데이터 추출
                for tab_name, tab_api_data in tabs_api_data.items():
                    if tab_api_data and 'response_data' in tab_api_data:
                        response_data = tab_api_data['response_data']
                        cards = response_data.get('cards', [])
                        
                        for card in cards:
                            if 'keyMetricCardData' in card:
                                tabs = card['keyMetricCardData'].get('keyMetricTabs', [])
                                for tab in tabs:
                                    primary_content = tab.get('primaryContent', {})
                                    metric = primary_content.get('metric', '')
                                    total = primary_content.get('total', 0)
                                    
                                    if metric == 'VIDEO_THUMBNAIL_IMPRESSIONS':
                                        impressions = total
                                    elif metric == 'VIDEO_THUMBNAIL_IMPRESSIONS_VTR':
                                        click_rate = total
                                    elif metric in ['EXTERNAL_VIEWS', 'VIEWS', 'VIEW_COUNT']:
                                        views = total
                                    elif metric in ['EXTERNAL_WATCH_TIME', 'WATCH_TIME']:
                                        watch_time_minutes = round(total / 1000 / 60, 1)
                                    elif metric in ['WATCH_TIME_MINUTES']:
                                        watch_time_minutes = total
                                    elif metric in ['AVERAGE_VIEW_DURATION', 'AVG_VIEW_DURATION', 'AVERAGE_VIEW_DURATION_SECONDS']:
                                        average_view_duration_seconds = total
                            
                            elif 'audienceRetentionHighlightsCardData' in card:
                                videos_data = card['audienceRetentionHighlightsCardData'].get('videosData', [])
                                for video_data in videos_data:
                                    if video_data.get('videoId') == video_id:
                                        metric_totals = video_data.get('metricTotals', {})
                                        
                                        avg_duration_millis = metric_totals.get('avgViewDurationMillis', 0)
                                        if avg_duration_millis:
                                            average_view_duration_seconds = round(avg_duration_millis / 1000)
                                        
                                        avg_percentage = metric_totals.get('avgPercentageWatched', 0)
                                        if avg_percentage:
                                            average_percentage_watched = round(avg_percentage * 100, 2)
                                        
                                        retention_views = metric_totals.get('views', 0)
                                        if retention_views and not views:
                                            views = retention_views
                                        break
                
                # 시간과 상태 변환 함수들
                def format_timestamp_korean(timestamp):
                    try:
                        if timestamp and str(timestamp) != 'N/A' and str(timestamp).isdigit():
                            dt = datetime.fromtimestamp(int(timestamp))
                            return dt.strftime('%Y년 %m월 %d일 %H시 %M분')
                        return 'N/A'
                    except:
                        return 'N/A'
                
                def format_duration(seconds):
                    try:
                        if seconds and str(seconds) != 'N/A' and str(seconds).isdigit():
                            seconds = int(seconds)
                            hours = seconds // 3600
                            minutes = (seconds % 3600) // 60
                            secs = seconds % 60
                            if hours > 0:
                                return f"{hours}:{minutes:02d}:{secs:02d}"
                            else:
                                return f"{minutes}:{secs:02d}"
                        return 'N/A'
                    except:
                        return 'N/A'
                
                # 공개 상태 변환
                privacy_mapping = {
                    'VIDEO_PRIVACY_PUBLIC': '공개',
                    'VIDEO_PRIVACY_PRIVATE': '비공개',
                    'VIDEO_PRIVACY_UNLISTED': '제한공개',
                    'VIDEO_PRIVACY_DRAFT': '임시보관함',
                    'VIDEO_PRIVACY_SCHEDULED': '예약 게시',
                    'VIDEO_STATUS_PROCESSED': '처리 완료',
                    'VIDEO_STATUS_UPLOADING': '업로드 중',
                    'VIDEO_STATUS_PROCESSING': '처리 중',
                    'VIDEO_STATUS_FAILED': '실패'
                }
                
                privacy_status = (
                    video_info.get('privacy') or 
                    video_info.get('status') or 
                    video_info.get('privacyStatus') or 
                    'Unknown'
                )
                privacy_korean = privacy_mapping.get(privacy_status, privacy_status)
                
                # 썸네일 URL 추출
                thumbnail_url = None
                if 'thumbnailDetails' in video_info and 'thumbnails' in video_info['thumbnailDetails']:
                    thumbnails = video_info['thumbnailDetails']['thumbnails']
                    if thumbnails and len(thumbnails) > 0:
                        thumbnail_url = thumbnails[-1].get('url')
                elif 'thumbnail_urls' in video_info:
                    urls = video_info['thumbnail_urls']
                    if urls and len(urls) > 0:
                        thumbnail_url = urls[-1]
                if not thumbnail_url and video_id != 'N/A':
                    thumbnail_url = f"https://img.youtube.com/vi/{video_id}/maxresdefault.jpg"
                
                # 데이터 행 작성 (썸네일을 첫 번째로)
                data = [
                    "썸네일 이미지",  # 첫 번째 열: 썸네일
                    video_id,
                    video_title[:50],
                    privacy_korean,
                    format_duration(video_info.get('lengthSeconds')),
                    format_timestamp_korean(video_info.get('timePublishedSeconds')),
                    int(video_info.get('public_viewCount', 0)) if video_info.get('public_viewCount') else 0,
                    int(impressions) if impressions else 0,
                    round(float(click_rate), 2) if click_rate else 0,  # 소수점으로 표시 (% 제거)
                    int(watch_time_minutes) if watch_time_minutes else 0,
                    int(average_view_duration_seconds) if average_view_duration_seconds else 0,
                    round(float(average_percentage_watched), 2) if average_percentage_watched else 0
                ]
                
                # 데이터 셀 작성 및 스타일 적용
                for col_num, value in enumerate(data, 1):
                    cell = worksheet.cell(row=row_idx, column=col_num, value=value)
                    cell.font = Font(name="맑은 고딕", size=10)
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                
                                        # 컬럼별 정렬 및 포맷 (썸네일이 첫 번째로 이동)
                    if col_num in [7, 8, 10, 11]:  # 숫자 컬럼들 (조회수, 노출수, 시청시간, 평균시청시간)
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        if isinstance(value, (int, float)) and value > 0:
                            cell.number_format = '#,##0'
                    elif col_num in [9, 12]:  # 소수점 컬럼들 (클릭률, 평균조회율)
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        if isinstance(value, (int, float)) and value > 0:
                            cell.number_format = '0.00'
                    elif col_num in [1, 2, 4, 5]:  # 중앙 정렬 (썸네일, ID, 공개상태, 길이)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # 썸네일 이미지 삽입 (1번째 열)
                thumbnail_cell = worksheet.cell(row=row_idx, column=1)
                if thumbnail_url:
                    try:
                        print(f"      📷 썸네일 다운로드: {thumbnail_url[:50]}...")
                        response = requests.get(thumbnail_url, timeout=15, headers={
                            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                        })
                        
                        if response.status_code == 200:
                            # BytesIO를 사용해서 메모리에서 처리
                            img_data = BytesIO(response.content)
                            img = Image(img_data)
                            
                            # 이미지 크기 조정 (더 크게)
                            img.width = 200
                            img.height = 150
                            
                            # 셀에 이미지 고정
                            img.anchor = f'A{row_idx}'
                            worksheet.add_image(img)
                            
                            # 행 높이를 더 크게 조정 (썸네일에 맞춰서)
                            worksheet.row_dimensions[row_idx].height = 120
                            
                            thumbnail_cell.value = "📷"
                            print(f"      ✅ 썸네일 삽입 완료")
                        else:
                            thumbnail_cell.value = f"❌ HTTP {response.status_code}"
                            print(f"      ❌ 썸네일 다운로드 실패: {response.status_code}")
                    except Exception as e:
                        thumbnail_cell.value = "❌ 오류"
                        print(f"      ❌ 썸네일 오류: {str(e)[:50]}")
                else:
                    thumbnail_cell.value = "❌ URL없음"
                

            
            # 컬럼 너비 설정 (썸네일이 첫 번째 열)
            column_widths = [30, 15, 40, 12, 8, 18, 12, 12, 10, 15, 15, 12]
            for col_num, width in enumerate(column_widths, 1):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
            
            # 행 높이 설정 (헤더는 작게, 데이터 행은 썸네일에 맞게)
            worksheet.row_dimensions[1].height = 25  # 헤더 행
            for row in range(2, len(analytics_data) + 2):
                worksheet.row_dimensions[row].height = 120  # 데이터 행들 (썸네일에 맞춰서)
            
            # 자동 필터 및 틀 고정
            worksheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(analytics_data) + 1}"
            worksheet.freeze_panes = "A2"
            
            # 파일 저장
            workbook.save(filename)
            
            print(f"✅ 📊 통합 애널리틱스 엑셀 파일이 생성되었습니다: {filename}")
            print(f"   • ✅ 기본 비디오 정보 + 애널리틱스 데이터 통합")
            print(f"   • ✅ 한국어 날짜 변환")
            print(f"   • ✅ 공개 상태 한글화")
            print(f"   • ✅ 비디오 URL 하이퍼링크")
            print(f"   • ✅ 썸네일 실제 이미지 삽입")
            print(f"   • ✅ 단일 시트만 존재")
                
        except Exception as e:
            print(f"❌ 간단 애널리틱스 엑셀 저장 오류: {e}")
            import traceback
            traceback.print_exc()
    

    


    def save_videos_to_excel(self, videos, filename):
        """비디오 데이터를 예쁘게 포맷된 엑셀 파일로 저장"""
        try:
            if not videos:
                print("❌ 저장할 비디오 데이터가 없습니다.")
                return

            # 필요한 임포트
            import requests
            from openpyxl.drawing.image import Image
            from io import BytesIO
            import tempfile
            import urllib.request

            print(f"📊 {len(videos)}개 비디오 데이터 처리 시작...")
            
            # 확장된 공개 상태 매핑 (모든 가능한 상태 포함)
            privacy_mapping = {
                'VIDEO_PRIVACY_PUBLIC': '공개',
                'VIDEO_PRIVACY_PRIVATE': '비공개', 
                'VIDEO_PRIVACY_UNLISTED': '제한공개',
                'VIDEO_PRIVACY_DRAFT': '임시보관함',
                'VIDEO_PRIVACY_SCHEDULED': '예약 게시',
                'VIDEO_STATUS_PROCESSED': '처리 완료',
                'VIDEO_STATUS_UPLOADING': '업로드 중',
                'VIDEO_STATUS_PROCESSING': '처리 중',
                'VIDEO_STATUS_FAILED': '실패',
                'PUBLIC': '공개',
                'PRIVATE': '비공개',
                'UNLISTED': '제한공개'
            }

            # 엑셀 워크북 생성 (단일 시트만)
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'YouTube Videos'
            
            # 헤더 설정
            headers = [
                '비디오 ID', '제목', '공개 상태', '길이', '생성일', '게시일',
                '조회수', '좋아요', '댓글수', '시청 시간 (분)', '구독자 증가',
                '상태', '비디오 URL', '썸네일', '설명'
            ]
            
            # 헤더 작성
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num, value=header)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # 데이터 처리
            for row_idx, video in enumerate(videos, 2):
                print(f"   📹 처리 중: {video.get('title', 'Unknown')[:30]}...")
                
                # 시간 변환 함수 (한국 시간으로)
                def format_timestamp_korean(timestamp):
                    try:
                        if timestamp and str(timestamp) != 'N/A' and str(timestamp).isdigit():
                            dt = datetime.fromtimestamp(int(timestamp))
                            return dt.strftime('%Y년 %m월 %d일 %H시 %M분')
                        return 'N/A'
                    except:
                        return 'N/A'
                
                # 길이를 시:분:초 형식으로 변환
                def format_duration(seconds):
                    try:
                        if seconds and str(seconds) != 'N/A' and str(seconds).isdigit():
                            seconds = int(seconds)
                            hours = seconds // 3600
                            minutes = (seconds % 3600) // 60
                            secs = seconds % 60
                            if hours > 0:
                                return f"{hours}:{minutes:02d}:{secs:02d}"
                            else:
                                return f"{minutes}:{secs:02d}"
                        return 'N/A'
                    except:
                        return 'N/A'
                
                # 숫자 안전 변환
                def safe_int(value, default=0):
                    try:
                        if value and str(value) != 'N/A':
                            return int(value)
                        return default
                    except:
                        return default
                
                # 비디오 ID와 URL
                video_id = video.get('videoId', 'N/A')
                
                # 공개 상태 변환 - 모든 가능한 필드 체크
                privacy_status = (
                    video.get('privacy') or 
                    video.get('status') or 
                    video.get('privacyStatus') or 
                    'Unknown'
                )
                privacy_korean = privacy_mapping.get(privacy_status, privacy_status)
                
                # 썸네일 URL 추출 - 여러 경로 시도
                thumbnail_url = None
                
                # 경로 1: thumbnailDetails.thumbnails
                if 'thumbnailDetails' in video and 'thumbnails' in video['thumbnailDetails']:
                    thumbnails = video['thumbnailDetails']['thumbnails']
                    if thumbnails and len(thumbnails) > 0:
                        # 가장 큰 해상도 선택
                        thumbnail_url = thumbnails[-1].get('url')
                
                # 경로 2: thumbnail_urls (이미 추출된 경우)
                if not thumbnail_url and 'thumbnail_urls' in video:
                    urls = video['thumbnail_urls']
                    if urls and len(urls) > 0:
                        thumbnail_url = urls[-1]
                
                # 경로 3: 직접 thumbnail 필드
                if not thumbnail_url:
                    thumbnail_url = video.get('thumbnail')
                
                # 경로 4: 비디오 ID로 YouTube 기본 썸네일 생성
                if not thumbnail_url and video_id != 'N/A':
                    thumbnail_url = f"https://img.youtube.com/vi/{video_id}/maxresdefault.jpg"
                
                # 데이터 행 작성
                data = [
                    video_id,
                    video.get('title', 'N/A'),
                    privacy_korean,
                    format_duration(video.get('lengthSeconds')),
                    format_timestamp_korean(video.get('timeCreatedSeconds')),
                    format_timestamp_korean(video.get('timePublishedSeconds')),
                    safe_int(video.get('public_viewCount', 0)),
                    safe_int(video.get('public_likeCount', 0)),
                    safe_int(video.get('public_commentCount', 0)),
                    safe_int(video.get('public_watchTimeMinutes', 0)),
                    safe_int(video.get('public_subscribersGained', 0)),
                    video.get('status', 'N/A'),
                    f"https://www.youtube.com/watch?v={video_id}",
                    "썸네일 이미지",  # 이미지가 들어갈 자리
                    video.get('description', 'N/A')[:200] if video.get('description') else 'N/A'
                ]
                
                # 데이터 셀 작성 및 스타일 적용
                for col_num, value in enumerate(data, 1):
                    cell = worksheet.cell(row=row_idx, column=col_num, value=value)
                    cell.font = Font(name="맑은 고딕", size=10)
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                    # 컬럼별 정렬
                    if col_num in [7, 8, 9, 10, 11]:  # 숫자 컬럼들
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        if isinstance(value, int) and value > 0:
                                cell.number_format = '#,##0'
                    elif col_num in [1, 3, 12]:  # 중앙 정렬
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # 비디오 URL을 하이퍼링크로 설정 (13번째 열)
                url_cell = worksheet.cell(row=row_idx, column=13)
                if video_id != 'N/A':
                    video_url = f"https://www.youtube.com/watch?v={video_id}"
                    url_cell.value = video_url
                    url_cell.hyperlink = video_url
                    url_cell.font = Font(name="맑은 고딕", size=10, color="0000FF", underline="single")
                    url_cell.alignment = Alignment(horizontal="left", vertical="center")
                    print(f"      🔗 하이퍼링크 설정: {video_url}")
                
                # 썸네일 이미지 삽입 (14번째 열)
                thumbnail_cell = worksheet.cell(row=row_idx, column=14)
                if thumbnail_url:
                    try:
                        print(f"      📷 썸네일 다운로드: {thumbnail_url[:50]}...")
                        response = requests.get(thumbnail_url, timeout=15, headers={
                            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                        })
                        
                        if response.status_code == 200:
                            # BytesIO를 사용해서 메모리에서 처리
                            img_data = BytesIO(response.content)
                            img = Image(img_data)
                            
                            # 이미지 크기 조정
                            img.width = 120
                            img.height = 90
                            
                            # 셀에 이미지 고정
                            img.anchor = f'{openpyxl.utils.get_column_letter(14)}{row_idx}'
                            worksheet.add_image(img)
                            
                            # 행 높이 조정
                            worksheet.row_dimensions[row_idx].height = 70
                            
                            thumbnail_cell.value = "📷 이미지"
                            print(f"      ✅ 썸네일 삽입 완료")
                        else:
                            thumbnail_cell.value = f"❌ HTTP {response.status_code}"
                            print(f"      ❌ 썸네일 다운로드 실패: {response.status_code}")
                    except Exception as e:
                        thumbnail_cell.value = "❌ 오류"
                        print(f"      ❌ 썸네일 오류: {str(e)[:50]}")
                else:
                    thumbnail_cell.value = "❌ URL없음"
            
            # 컬럼 너비 설정
            column_widths = [15, 50, 12, 10, 22, 22, 12, 10, 10, 15, 12, 12, 50, 25, 60]
            for col_num, width in enumerate(column_widths, 1):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
                
            # 자동 필터 및 틀 고정
            worksheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(videos) + 1}"
            worksheet.freeze_panes = "A2"
                
            # 파일 저장
            workbook.save(filename)
            
            print(f"✅ 📊 완전히 새로운 엑셀 파일이 생성되었습니다: {filename}")
            print(f"   • ✅ 유닉스 타임스탬프 → 한국어 날짜")
            print(f"   • ✅ 모든 공개 상태 한글화")
            print(f"   • ✅ 비디오 URL 하이퍼링크")
            print(f"   • ✅ 썸네일 실제 이미지 삽입")
            print(f"   • ✅ 단일 시트만 (상세보기 제거)")
            
        except Exception as e:
            print(f"❌ 엑셀 저장 오류: {e}")
            import traceback
            traceback.print_exc()

    def navigate_to_video_analytics_tab(self, video_id, tab_config):
        """특정 비디오의 특정 애널리틱스 탭으로 이동"""
        try:
            if not self.channel_id:
                print("❌ 채널 ID가 없습니다.")
                return False
            
            # 애널리틱스 탭 URL 생성
            analytics_url = f"https://studio.youtube.com/video/{video_id}/analytics/{tab_config['url_suffix']}"
            
            # 활성 탭에서 URL 변경
            response = requests.get(f'http://localhost:{self.chrome_port}/json')
            tabs = response.json()
            
            studio_tab = None
            for tab in tabs:
                if 'studio.youtube.com' in tab.get('url', ''):
                    studio_tab = tab
                    break
            
            if studio_tab:
                # JavaScript로 페이지 이동
                ws_url = studio_tab['webSocketDebuggerUrl']
                ws = websocket.create_connection(ws_url)
                
                ws.send(json.dumps({
                    "id": 1,
                    "method": "Runtime.evaluate",
                    "params": {
                        "expression": f"window.location.href = '{analytics_url}'"
                    }
                }))
                
                ws.close()
                print(f"✅ {tab_config['name']} 탭으로 이동: {analytics_url}")
                print(f"   📊 수집 예정: {tab_config['description']}")
                time.sleep(8)  # 페이지 로딩 대기
                return True
            
            return False
            
        except Exception as e:
            print(f"❌ {tab_config['name']} 탭 이동 오류: {e}")
            return False
    
    def start_multi_tab_analytics_collection(self, video_id):
        """한 비디오에 대해 모든 애널리틱스 탭을 순차적으로 수집"""
        try:
            print(f"🎯 다중 탭 애널리틱스 수집 시작 (비디오 ID: {video_id})")
            
            # 각 비디오마다 안전하게 초기화
            with self.tab_lock:
                self.current_tab_index = 0  # 반드시 0으로 리셋!
                self.collected_analytics_data = {}  # 이전 데이터 클리어
            
            # 첫 번째 비디오인지 확인 (이미 수집된 데이터 개수로 판단)
            is_first_video = len(self.video_analytics_data) == 0
            
            if is_first_video:
                # 첫 번째 비디오: 브라우저 이동해서 요청 캡처
                print(f"🆕 첫 번째 비디오: 브라우저 탭 이동으로 요청 캡처")
                
                if self.analytics_tabs and len(self.analytics_tabs) > 0:
                    first_tab = self.analytics_tabs[0]
                    print(f"🚀 1단계: {first_tab['name']} 탭 수집 시작")
                    threading.Thread(
                        target=self.navigate_to_video_analytics_tab,
                        args=(video_id, first_tab),
                        daemon=True
                    ).start()
                else:
                    print("❌ 애널리틱스 탭 목록이 비어있습니다!")
            else:
                # 두 번째 비디오부터: 캡처된 요청 사용해서 바로 데이터 수집
                print(f"🚀 두 번째 이후 비디오: 캐시된 요청으로 직접 API 호출")
                self.process_video_with_captured_requests(video_id)
                
        except Exception as e:
            print(f"❌ 다중 탭 수집 시작 오류: {e}")
            import traceback
            traceback.print_exc()
            # 오류 시 다음 비디오로 진행
            self.proceed_to_next_video()
    
    def process_video_with_captured_requests(self, video_id):
        """캡처된 요청들을 사용해서 비디오 데이터를 바로 수집"""
        try:
            print(f"🚀 캐시된 요청으로 {video_id} 비디오 데이터 수집 중...")
            
            # 캐시된 요청 확인
            if not hasattr(self, 'captured_reach_viewers_request') or not hasattr(self, 'captured_interest_viewers_request'):
                print(f"❌ 캐시된 요청이 없습니다. 첫 번째 비디오가 제대로 수집되지 않았을 수 있습니다.")
                # 다음 비디오로 바로 진행
                self.proceed_to_next_video()
                return
            
            print(f"✅ 캐시된 요청 발견. API 호출 시작...")
                    
            # 수집된 탭별 실제 API 응답 데이터 저장
            collected_api_responses = {}
            
            # reach_viewers 탭 처리
            print(f"📡 reach_viewers 탭 데이터 수집 중...")
            try:
                api_response = self.replay_analytics_request(self.captured_reach_viewers_request, video_id)
                if api_response:
                    collected_api_responses['reach_viewers'] = {
                        'api_type': 'get_screen',
                        'response_data': api_response,
                        'tab_config': self.analytics_tabs[0]  # reach_viewers는 첫 번째 탭
                    }
                    print(f"   ✅ reach_viewers: 성공")
                    self.extract_metrics_from_get_screen_response_immediate(api_response, video_id)
                else:
                    print(f"   ❌ reach_viewers: 실패")
            except Exception as e:
                print(f"   ❌ reach_viewers 오류: {e}")
            
            # interest_viewers 탭 처리  
            print(f"📡 interest_viewers 탭 데이터 수집 중...")
            try:
                api_response = self.replay_analytics_request(self.captured_interest_viewers_request, video_id)
                if api_response:
                    collected_api_responses['interest_viewers'] = {
                        'api_type': 'get_cards',
                        'response_data': api_response,
                        'tab_config': self.analytics_tabs[1]  # interest_viewers는 두 번째 탭
                    }
                    print(f"   ✅ interest_viewers: 성공")
                    self.extract_metrics_from_get_cards_response_immediate(api_response, video_id)
                else:
                    print(f"   ❌ interest_viewers: 실패")
            except Exception as e:
                print(f"   ❌ interest_viewers 오류: {e}")
            
            # 기본 비디오 정보 찾기
            basic_video_info = None
            for video in self.collected_data:
                if video.get('videoId') == video_id:
                    basic_video_info = video
                    break
            
            # 중복 체크: 이미 수집된 비디오인지 확인
            already_exists = any(item.get('video_id') == video_id for item in self.video_analytics_data)
            
            if not already_exists:
                # 종합된 데이터 저장
                combined_data = {
                    'video_id': video_id,
                    'video_title': basic_video_info.get('title') if basic_video_info else 'Unknown',
                    'collected_at': datetime.now().isoformat(),
                    'basic_video_info': basic_video_info,
                    'tabs_data': {},  # 캐시된 요청 사용이므로 빈 값
                    'analytics_data': collected_api_responses  # 실제 API 응답 데이터
                }
                
                self.video_analytics_data.append(combined_data)
                print(f"✅ 새로운 비디오 데이터 추가됨: {video_id}")
            else:
                print(f"⚠️ 비디오 {video_id}는 이미 수집되었습니다. 중복 추가 방지.")
            
            print(f"✅ 비디오 {video_id} 데이터 수집 완료 ({len(collected_api_responses)}/{len(self.analytics_tabs)} 탭)")
            print(f"💾 현재까지 수집된 비디오: {len(self.video_analytics_data)}개")
            
            # 다음 비디오로 진행
            self.proceed_to_next_video()
            
        except Exception as e:
            print(f"❌ 캐시된 요청으로 비디오 처리 오류: {e}")
            import traceback
            traceback.print_exc()
            # 오류가 나도 다음 비디오로 진행
            self.proceed_to_next_video()
    
    def proceed_to_next_tab(self, current_video_id):
        """다음 탭으로 이동하거나 다음 비디오로 진행"""
        try:
            with self.tab_lock:  # 스레드 안전성을 위한 락
                print(f"🔍 현재 탭 인덱스: {self.current_tab_index}")
                print(f"🔍 전체 탭 수: {len(self.analytics_tabs)}")
                
                # 다음 인덱스를 미리 계산해서 범위 체크
                next_index = self.current_tab_index + 1
                print(f"🔍 계산된 다음 탭 인덱스: {next_index}")
                
                if next_index < len(self.analytics_tabs):
                    # 안전하게 인덱스 증가
                    self.current_tab_index = next_index
                    
                    # 다음 탭으로 이동
                    next_tab = self.analytics_tabs[self.current_tab_index]
                    print(f"🚀 {self.current_tab_index + 1}단계: {next_tab['name']} 탭 수집 시작")
                    threading.Thread(
                        target=self.navigate_to_video_analytics_tab,
                        args=(current_video_id, next_tab),
                        daemon=True
                    ).start()
                else:
                    # 현재 비디오의 모든 탭 수집 완료
                    print(f"🎉 비디오 {current_video_id}의 모든 탭 수집 완료!")
                    print(f"   수집된 탭들: {list(self.collected_analytics_data.keys())}")
                    self.finalize_video_analytics(current_video_id)
                    
                    # 다음 비디오로 진행하거나 전체 수집 완료
                    self.proceed_to_next_video()
                
        except Exception as e:
            print(f"❌ 다음 탭 진행 오류: {e}")
            import traceback
            traceback.print_exc()
    
    def finalize_video_analytics(self, video_id):
        """한 비디오의 모든 탭 데이터를 종합하고 실제 API 응답 데이터 수집"""
        try:
            print(f"📊 비디오 {video_id} 데이터 종합 시작...")
            
            # 수집된 탭별 실제 API 응답 데이터 가져오기
            collected_api_responses = {}
            
            for tab_name, tab_data in self.collected_analytics_data.items():
                if tab_data and 'captured_request' in tab_data:
                    print(f"🔄 {tab_name} 탭 API 응답 수집 중...")
                    
                    captured_request = tab_data['captured_request']
                    api_type = tab_data['api_type']
                    
                    # 캡처된 요청을 복제해서 실제 응답 데이터 가져오기
                    api_response = self.replay_analytics_request(captured_request, video_id)
                    
                    if api_response:
                        collected_api_responses[tab_name] = {
                            'api_type': api_type,
                            'response_data': api_response,
                            'tab_config': tab_data['tab_config']
                        }
                        print(f"   ✅ {tab_name}: API 응답 수집 완료")
                        
                        # CTR과 노출수 즉시 추출 및 출력
                        if api_type == 'get_screen':
                            self.extract_metrics_from_get_screen_response_immediate(api_response, video_id)
                        elif api_type == 'get_cards':
                            self.extract_metrics_from_get_cards_response_immediate(api_response, video_id)
                    else:
                        print(f"   ❌ {tab_name}: API 응답 수집 실패")
                else:
                    print(f"   ⚠️ {tab_name}: 캡처된 요청 없음")
            
            # 기본 비디오 정보 찾기
            basic_video_info = None
            for video in self.collected_data:
                if video.get('videoId') == video_id:
                    basic_video_info = video
                    break
            
            # 중복 체크: 이미 수집된 비디오인지 확인
            already_exists = any(item.get('video_id') == video_id for item in self.video_analytics_data)
            
            if not already_exists:
                # 종합된 데이터 저장
                combined_data = {
                    'video_id': video_id,
                    'video_title': basic_video_info.get('title') if basic_video_info else 'Unknown',
                    'collected_at': datetime.now().isoformat(),
                    'basic_video_info': basic_video_info,
                    'tabs_data': self.collected_analytics_data.copy(),
                    'analytics_data': collected_api_responses  # 실제 API 응답 데이터
                }
                
                self.video_analytics_data.append(combined_data)
                print(f"✅ 새로운 비디오 데이터 추가됨: {video_id}")
            else:
                print(f"⚠️ 비디오 {video_id}는 이미 수집되었습니다. 중복 추가 방지.")
            
            print(f"📊 비디오 {video_id} 데이터 종합 완료:")
            for tab_name, tab_data in self.collected_analytics_data.items():
                if tab_data:
                    print(f"   ✅ {tab_name}: 데이터 수집됨")
                else:
                    print(f"   ❌ {tab_name}: 데이터 없음")
            
            # 첫 번째 비디오라면 캡처된 요청들을 저장해서 다른 비디오들이 사용할 수 있게 함
            if len(self.video_analytics_data) == 1:
                print(f"💾 첫 번째 비디오의 캡처된 요청들을 저장 중...")
                for tab_name, tab_data in self.collected_analytics_data.items():
                    if tab_data and 'captured_request' in tab_data:
                        if tab_name == 'reach_viewers':
                            self.captured_reach_viewers_request = tab_data['captured_request']
                            print(f"   ✅ reach_viewers 요청 저장됨")
                        elif tab_name == 'interest_viewers':
                            self.captured_interest_viewers_request = tab_data['captured_request']
                            print(f"   ✅ interest_viewers 요청 저장됨")
            
            print(f"💾 현재까지 수집된 비디오: {len(self.video_analytics_data)}개")
            print(f"🔄 다음 비디오 처리 또는 전체 완료 대기 중...")
            
        except Exception as e:
            print(f"❌ 비디오 데이터 종합 오류: {e}")
            import traceback
            traceback.print_exc()
    
    def proceed_to_next_video(self):
        """다음 비디오 처리 또는 전체 수집 완료"""
        try:
            print(f"🔄 다음 비디오로 진행 중... (현재 수집: {len(self.video_analytics_data)}개)")
            
            # 다음 비디오 찾기
            if len(self.video_analytics_data) < len(self.collected_data):
                next_video_index = len(self.video_analytics_data)
                next_video = self.collected_data[next_video_index]
                next_video_id = next_video.get('videoId')
                next_video_title = next_video.get('title', '제목없음')[:30]
                
                print(f"📹 [{next_video_index + 1}/{len(self.collected_data)}] {next_video_title} (ID: {next_video_id})")
                
                # 다음 비디오의 다중 탭 수집 시작 (짧은 지연 후)
                def delayed_start():
                    time.sleep(2)  # 2초 대기로 안정성 확보
                    self.start_multi_tab_analytics_collection(next_video_id)
                
                threading.Thread(target=delayed_start, daemon=True).start()
                
            else:
                # 모든 비디오 처리 완료
                print(f"🎉 모든 {len(self.collected_data)}개 비디오의 수집 완료!")
                print(f"📊 총 수집된 비디오: {len(self.video_analytics_data)}개")
                
                if self.video_analytics_data:
                    print(f"💾 엑셀 파일로 저장합니다...")
                self.save_analytics_data(self.video_analytics_data)
                
                # 모니터링 중단
                print(f"🏁 모든 데이터 수집 완료! 모니터링을 중단합니다.")
                self.monitoring = False
            
        except Exception as e:
            print(f"❌ 다음 비디오 진행 오류: {e}")
            import traceback
            traceback.print_exc()
            # 오류 시에도 모니터링 중단
            print(f"🛑 오류로 인해 모니터링을 중단합니다.")
            self.monitoring = False
    
    def replay_analytics_request(self, captured_request, video_id):
        """캡처된 애널리틱스 요청을 복제해서 API 응답 가져오기"""
        try:
            url = captured_request['url']
            method = captured_request['method']
            headers = captured_request['headers'].copy()
            post_data = captured_request['postData']
            
            print(f"   📡 API 요청 시작... (타임아웃: 15초)")
            
            # 비디오 ID 교체 (필요한 경우)
            if post_data:
                try:
                    payload = json.loads(post_data)
                    modified_payload = self.replace_video_id_in_payload(payload, video_id)
                    if modified_payload:
                        post_data = json.dumps(modified_payload, separators=(',', ':'))
                        print(f"   🔄 비디오 ID 교체 완료")
                except:
                    print(f"   ⚠️ 비디오 ID 교체 실패 - 원본 사용")
            
            # requests 세션 생성
            session = requests.Session()
            
            # 쿠키 설정
            if 'Cookie' in headers:
                cookie_str = headers['Cookie']
                cookies = {}
                for item in cookie_str.split(';'):
                    if '=' in item:
                        key, value = item.strip().split('=', 1)
                        cookies[key] = value
                session.cookies.update(cookies)
                print(f"   🍪 쿠키 설정 완료 ({len(cookies)}개)")
            
            # Content-Length 제거
            headers.pop('Content-Length', None)
            
            # 요청 전송 (더 짧은 타임아웃)
            if method.upper() == 'POST' and post_data:
                try:
                    json_data = json.loads(post_data)
                    response = session.post(url, headers=headers, json=json_data, timeout=15)
                except json.JSONDecodeError:
                    response = session.post(url, headers=headers, data=post_data, timeout=15)
            else:
                response = session.get(url, headers=headers, timeout=15)
            
            print(f"   📨 응답 수신: {response.status_code}")
            
            if response.status_code == 200:
                try:
                    return response.json()
                except json.JSONDecodeError:
                    print(f"   ❌ JSON 파싱 실패")
                    return None
            else:
                print(f"   ❌ HTTP 에러: {response.status_code}")
                return None
                
        except requests.Timeout:
            print(f"   ⏰ API 요청 타임아웃 (15초)")
            return None
        except Exception as e:
            print(f"   ❌ API 요청 오류: {e}")
            return None
    
    def extract_metrics_from_get_cards_response_immediate(self, cards_data, video_id):
        """get_cards API 응답에서 즉시 메트릭 추출 및 출력"""
        try:
            print(f"📊 [{video_id}] get_cards 데이터에서 조회수/시청시간 추출:")
            
            if 'cards' in cards_data:
                cards = cards_data['cards']
                for card in cards:
                    if 'keyMetricCardData' in card:
                        tabs = card['keyMetricCardData'].get('keyMetricTabs', [])
                        for tab in tabs:
                            primary_content = tab.get('primaryContent', {})
                            metric = primary_content.get('metric', '')
                            total = primary_content.get('total', 0)
                            
                            if metric in ['EXTERNAL_VIEWS', 'VIEWS']:
                                print(f"   📊 조회수: {total:,}")
                            elif metric in ['EXTERNAL_WATCH_TIME', 'WATCH_TIME']:
                                # 시청시간 밀리초→분 변환
                                print(f"   🔍 시청시간 원본: 메트릭='{metric}', 값={total:,}")
                                minutes = round(total / 1000 / 60, 1)
                                print(f"   📊 시청시간(밀리초→분): {total:,}ms = {minutes:.1f}분")
                            elif metric in ['WATCH_TIME_MINUTES']:
                                print(f"   🔍 시청시간 원본: 메트릭='{metric}', 값={total:,}")
                                print(f"   📊 시청시간(분단위): {total:,}분")
                            elif metric in ['AVERAGE_VIEW_DURATION', 'AVG_VIEW_DURATION', 'AVERAGE_VIEW_DURATION_SECONDS', 'VIEW_DURATION_AVG']:
                                print(f"   📊 평균 시청시간: {total:,}초")
                            elif metric in ['SUBSCRIBERS_NET_CHANGE']:
                                print(f"   📊 구독자 증감: {total:+,}")
                            elif metric:
                                print(f"   📊 {metric}: {total:,}")
                    
                    # audienceRetentionHighlightsCardData에서 평균 시청시간과 평균 조회율 추출
                    elif 'audienceRetentionHighlightsCardData' in card:
                        videos_data = card['audienceRetentionHighlightsCardData'].get('videosData', [])
                        for video_data in videos_data:
                            if video_data.get('videoId') == video_id:
                                metric_totals = video_data.get('metricTotals', {})
                                
                                # 평균 시청시간 (밀리초 → 초)
                                avg_duration_millis = metric_totals.get('avgViewDurationMillis', 0)
                                if avg_duration_millis:
                                    avg_duration_seconds = round(avg_duration_millis / 1000)
                                    print(f"   📊 평균 시청시간: {avg_duration_millis:,}ms = {avg_duration_seconds:,}초")
                                
                                # 평균 조회율 (소수 → 퍼센트)
                                avg_percentage = metric_totals.get('avgPercentageWatched', 0)
                                if avg_percentage:
                                    avg_percentage_display = round(avg_percentage * 100, 2)
                                    print(f"   📊 평균 조회율: {avg_percentage:.4f} = {avg_percentage_display}%")
                                
                                break
            
        except Exception as e:
            print(f"❌ get_cards 메트릭 추출 오류: {e}")
    
    def extract_metrics_from_get_screen_response_immediate(self, screen_data, video_id):
        """get_screen API 응답에서 즉시 메트릭 추출 및 출력"""
        try:
            print(f"📊 [{video_id}] get_screen 데이터에서 CTR/노출수 추출:")
            
            if 'cards' in screen_data:
                cards = screen_data['cards']
                for card in cards:
                    if 'keyMetricCardData' in card:
                        tabs = card['keyMetricCardData'].get('keyMetricTabs', [])
                        for tab in tabs:
                            primary_content = tab.get('primaryContent', {})
                            metric = primary_content.get('metric', '')
                            total = primary_content.get('total', 0)
                            
                            if metric == 'VIDEO_THUMBNAIL_IMPRESSIONS':
                                print(f"   📊 노출수: {total:,}")
                            elif metric == 'VIDEO_THUMBNAIL_IMPRESSIONS_VTR':
                                print(f"   📊 클릭률: {total}%")
                            elif metric in ['AVERAGE_VIEW_DURATION', 'AVG_VIEW_DURATION', 'AVERAGE_VIEW_DURATION_SECONDS', 'VIEW_DURATION_AVG']:
                                print(f"   📊 평균 시청시간: {total:,}초")
                            elif metric:
                                print(f"   📊 {metric}: {total:,}")
                    
                    # audienceRetentionHighlightsCardData에서 평균 시청시간과 평균 조회율 추출
                    elif 'audienceRetentionHighlightsCardData' in card:
                        videos_data = card['audienceRetentionHighlightsCardData'].get('videosData', [])
                        for video_data in videos_data:
                            if video_data.get('videoId') == video_id:
                                metric_totals = video_data.get('metricTotals', {})
                                
                                # 평균 시청시간 (밀리초 → 초)
                                avg_duration_millis = metric_totals.get('avgViewDurationMillis', 0)
                                if avg_duration_millis:
                                    avg_duration_seconds = round(avg_duration_millis / 1000)
                                    print(f"   📊 평균 시청시간: {avg_duration_millis:,}ms = {avg_duration_seconds:,}초")
                                
                                # 평균 조회율 (소수 → 퍼센트)
                                avg_percentage = metric_totals.get('avgPercentageWatched', 0)
                                if avg_percentage:
                                    avg_percentage_display = round(avg_percentage * 100, 2)
                                    print(f"   📊 평균 조회율: {avg_percentage:.4f} = {avg_percentage_display}%")
                                
                                break
            
        except Exception as e:
            print(f"❌ get_screen 메트릭 추출 오류: {e}")
    
    def extract_metrics_from_get_cards_response(self, cards_data):
        """get_cards API 응답에서 메트릭 추출"""
        try:
            extracted_metrics = {}
            
            if not cards_data or 'cards' not in cards_data:
                print("⚠️ cards 데이터가 없습니다.")
                return extracted_metrics
            
            cards = cards_data['cards']
            print(f"🔍 {len(cards)}개 카드에서 데이터 추출 중...")
            
            for i, card in enumerate(cards):
                print(f"   카드 {i+1}: {list(card.keys())}")
                
                # keyMetricCardData에서 주요 지표 추출
                if 'keyMetricCardData' in card:
                    tabs = card['keyMetricCardData'].get('keyMetricTabs', [])
                    for tab in tabs:
                        primary_content = tab.get('primaryContent', {})
                        metric = primary_content.get('metric', '')
                        total = primary_content.get('total', 0)
                        
                        if metric and total is not None:
                            print(f"     📊 {metric}: {total:,}")
                            extracted_metrics[metric] = total
                            
                            # 시계열 데이터도 추출 (타임스탬프 변환)
                            main_series = primary_content.get('mainSeries', {})
                            datums = main_series.get('datums', [])
                            if datums:
                                converted_datums = []
                                for datum in datums:
                                    x = datum.get('x')  # 유닉스 타임스탬프
                                    y = datum.get('y')  # 값
                                    converted_datums.append({
                                        'timestamp': self.convert_unix_timestamp(x),
                                        'unix_timestamp': x,
                                        'value': y
                                    })
                                extracted_metrics[f'{metric}_timeseries'] = converted_datums
                
                # personalizedHeaderCardData에서 제목 정보 추출
                if 'personalizedHeaderCardData' in card:
                    title = card['personalizedHeaderCardData'].get('title', '')
                    if title:
                        print(f"     📝 카드 제목: {title}")
                        extracted_metrics['header_title'] = title
            
            print(f"✅ 총 {len(extracted_metrics)}개 메트릭 추출 완료")
            return extracted_metrics
            
        except Exception as e:
            print(f"❌ get_cards 데이터 추출 오류: {e}")
            return {}
    
    def extract_metrics_from_get_screen_response(self, screen_data):
        """get_screen API 응답에서 메트릭 추출"""
        try:
            extracted_metrics = {}
            
            if not screen_data:
                print("⚠️ screen 데이터가 없습니다.")
                return extracted_metrics
            
            print(f"🔍 get_screen 데이터에서 메트릭 추출 중...")
            print(f"   응답 키들: {list(screen_data.keys())}")
            
            # get_screen 응답 구조에 따라 데이터 추출 로직 구현
            # (실제 응답 구조를 보고 나중에 상세 구현)
            
            return extracted_metrics
            
        except Exception as e:
            print(f"❌ get_screen 데이터 추출 오류: {e}")
            return {}

def main():
    print("YouTube Studio 다중 탭 애널리틱스 수집 시스템")
    print("=" * 80)
    print("🎯 실제 네트워크 요청을 감지하고 그대로 복제하여 완전한 데이터를 수집합니다!")
    print("\n📋 사용 방법:")
    print("   🔸 1. Chrome 브라우저가 자동으로 열림")
    print("   🔸 2. 원하는 Google 계정으로 로그인")
    print("   🔸 3. YouTube Studio로 이동하여 채널 선택")
    print("   🔸 4. 터미널에서 엔터를 눌러 수집 시작")
    print("\n📋 수집 과정:")
    print("   🔸 1단계: 비디오 목록 수집")
    print("     • list_creator_videos API 요청 감지 및 복제")
    print("     • 채널의 모든 비디오 기본 정보 수집")
    print("   🔸 2단계: 각 비디오 다중 탭 애널리틱스 수집")  
    print("     • tab-reach_viewers → get_screen API (노출수, CTR)")
    print("     • tab-interest_viewers → get_cards API (조회수, 시청시간, 구독자)")
    print("     • 각 탭을 순차적으로 방문하여 모든 데이터 수집")
    print("     • 유닉스 타임스탬프를 사람이 읽을 수 있는 형식으로 변환")
    print("\n✨ 특징:")
    print("   • 계정 변경 가능 - 수동 로그인으로 원하는 계정 선택")
    print("   • 다중 탭 순차 수집으로 완전한 애널리틱스 데이터 확보")
    print("   • 하드코딩된 페이로드 없음")
    print("   • 브라우저의 실제 요청을 그대로 복제")
    print("   • YouTube API 변경에도 대응 가능")
    print("   • 자동으로 모든 데이터를 JSON으로 저장")
    print("   • 성공한 요청만 복제하여 높은 성공률")
    
    monitor = YouTubeStudioMonitor()
    
    try:
        # 모니터링 시작 (충분한 시간 제공)
        monitor.start_monitoring(300)  # 5분으로 증가 (2단계 수집)
        
    except KeyboardInterrupt:
        print("\n사용자에 의해 중단되었습니다.")
    except Exception as e:
        print(f"오류 발생: {e}")
    finally:
        monitor.cleanup()

if __name__ == "__main__":
    main()
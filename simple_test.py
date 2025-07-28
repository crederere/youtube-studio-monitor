#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO
import requests
from datetime import datetime
import time

def create_simple_excel():
    """사용자 요청에 맞는 단순화된 엑셀 생성"""
    
    # 워크북 생성
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'YouTube Analytics'
    
    # 헤더 설정 (사용자 요청에 맞게 단순화)
    headers = [
        '썸네일', '비디오 ID', '제목', '공개상태', '길이', '생성일', '게시일',
        '조회수', '노출수', '클릭률 (%)', 
        '시청시간 (분)', '평균 시청시간 (초)', 
        '평균 조회율 (%)', '비디오 URL'
    ]
    
    # 헤더 작성
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    # 테스트 데이터
    test_videos = [
        {
            'videoId': 'RLe5T6Fn3YQ',
            'title': '단순화된 구조 테스트',
            'privacy': 'VIDEO_PRIVACY_PUBLIC',
            'lengthSeconds': '720',
            'timeCreatedSeconds': str(int(time.time() - 86400 * 30)),
            'timePublishedSeconds': str(int(time.time() - 86400 * 25)),
            'public_viewCount': '1500',
            'impressions': 2500,
            'click_rate': 8.5,
            'watch_time_minutes': 450,
            'avg_duration_seconds': 45,
            'avg_percentage': 65.2
        },
        {
            'videoId': 'oRWchhS6OqM',
            'title': '썸네일 첫 번째 열 테스트',
            'privacy': 'VIDEO_PRIVACY_PUBLIC',
            'lengthSeconds': '900',
            'timeCreatedSeconds': str(int(time.time() - 86400 * 7)),
            'timePublishedSeconds': str(int(time.time() - 86400 * 5)),
            'public_viewCount': '2300',
            'impressions': 3800,
            'click_rate': 12.3,
            'watch_time_minutes': 780,
            'avg_duration_seconds': 67,
            'avg_percentage': 78.5
        }
    ]
    
    # 공개 상태 변환
    privacy_mapping = {
        'VIDEO_PRIVACY_PUBLIC': '공개',
        'VIDEO_PRIVACY_PRIVATE': '비공개',
        'VIDEO_PRIVACY_UNLISTED': '제한공개'
    }
    
    # 시간 변환 함수
    def format_timestamp_korean(timestamp):
        try:
            if timestamp and str(timestamp).isdigit():
                dt = datetime.fromtimestamp(int(timestamp))
                return dt.strftime('%Y년 %m월 %d일 %H시 %M분')
            return 'N/A'
        except:
            return 'N/A'
    
    def format_duration(seconds):
        try:
            if seconds and str(seconds).isdigit():
                seconds = int(seconds)
                minutes = seconds // 60
                secs = seconds % 60
                return f"{minutes}:{secs:02d}"
            return 'N/A'
        except:
            return 'N/A'
    
    # 데이터 행 작성
    for row_idx, video in enumerate(test_videos, 2):
        video_id = video['videoId']
        privacy_korean = privacy_mapping.get(video['privacy'], video['privacy'])
        
        # 단순화된 데이터 (사용자 요청에 맞게)
        data = [
            "📷 썸네일",  # 썸네일 (첫 번째 열)
            video_id,
            video['title'],
            privacy_korean,
            format_duration(video['lengthSeconds']),
            format_timestamp_korean(video['timeCreatedSeconds']),
            format_timestamp_korean(video['timePublishedSeconds']),
            int(video['public_viewCount']),  # 조회수 (애널리틱스 조회수 없음)
            video['impressions'],  # 노출수
            video['click_rate'],  # 클릭률
            video['watch_time_minutes'],  # 시청시간
            video['avg_duration_seconds'],  # 평균 시청시간
            video['avg_percentage'],  # 평균 조회율
            f"https://www.youtube.com/watch?v={video_id}"  # 비디오 URL
        ]
        
        # 데이터 셀 작성
        for col_num, value in enumerate(data, 1):
            cell = worksheet.cell(row=row_idx, column=col_num, value=value)
            cell.font = Font(name="맑은 고딕", size=10)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # 컬럼별 정렬
            if col_num in [8, 9, 11, 12]:  # 숫자 컬럼들
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(value, (int, float)) and value > 0:
                    cell.number_format = '#,##0'
            elif col_num in [10, 13]:  # 퍼센트 컬럼들
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(value, (int, float)) and value > 0:
                    cell.number_format = '0.00'
            elif col_num in [1, 2, 4, 5]:  # 중앙 정렬
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 썸네일 이미지 삽입 (1번째 열)
        thumbnail_cell = worksheet.cell(row=row_idx, column=1)
        thumbnail_url = f"https://img.youtube.com/vi/{video_id}/maxresdefault.jpg"
        
        try:
            print(f"📷 썸네일 다운로드: {video_id}")
            response = requests.get(thumbnail_url, timeout=15, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            })
            
            if response.status_code == 200:
                img_data = BytesIO(response.content)
                img = Image(img_data)
                
                # 이미지 크기 조정 (크게)
                img.width = 160
                img.height = 120
                
                # 셀에 이미지 고정
                img.anchor = f'A{row_idx}'
                worksheet.add_image(img)
                
                # 행 높이를 크게 조정
                worksheet.row_dimensions[row_idx].height = 100
                
                thumbnail_cell.value = "📷"
                print(f"   ✅ 썸네일 삽입 완료")
            else:
                thumbnail_cell.value = f"❌ HTTP {response.status_code}"
        except Exception as e:
            thumbnail_cell.value = "❌ 오류"
            print(f"   ❌ 썸네일 오류: {str(e)[:50]}")
        
        # 비디오 URL을 하이퍼링크로 설정 (14번째 열)
        url_cell = worksheet.cell(row=row_idx, column=14)
        video_url = f"https://www.youtube.com/watch?v={video_id}"
        url_cell.value = video_url
        url_cell.hyperlink = video_url
        url_cell.font = Font(name="맑은 고딕", size=10, color="0000FF", underline="single")
        url_cell.alignment = Alignment(horizontal="left", vertical="center")
        print(f"   🔗 하이퍼링크 설정: {video_url}")
    
    # 컬럼 너비 설정
    column_widths = [25, 15, 40, 12, 8, 18, 18, 12, 12, 10, 15, 15, 12, 45]
    for col_num, width in enumerate(column_widths, 1):
        column_letter = openpyxl.utils.get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = width
    
    # 자동 필터 및 틀 고정
    worksheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(test_videos) + 1}"
    worksheet.freeze_panes = "A2"
    
    # 파일 저장
    filename = "SIMPLE_youtube_analytics.xlsx"
    workbook.save(filename)
    
    print(f"✅ 단순화된 엑셀 파일 생성 완료: {filename}")
    print(f"📋 구조:")
    print(f"   • 썸네일: 첫 번째 열, 크기 160x120, 행높이 100")
    print(f"   • 조회수: 애널리틱스 조회수 제거, 기본 조회수만")
    print(f"   • 좋아요/댓글수: 제거됨")
    print(f"   • 총 {len(headers)}개 컬럼")
    
    return filename

if __name__ == "__main__":
    create_simple_excel() 
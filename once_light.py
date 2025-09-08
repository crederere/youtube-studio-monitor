import json
import time
import requests
from datetime import datetime
import websocket
import threading
from urllib.parse import urlparse, parse_qs
import re
import os
import subprocess
import psutil
import webbrowser
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO
import tempfile
import urllib.request
import platform
import shutil

class YouTubeStudioMonitor:
    def __init__(self, chrome_port=9222):
        self.chrome_port = chrome_port
        self.ws = None
        self.captured_request = None  # 캡처된 원본 요청 저장
        self.captured_analytics_request = None  # 캡처된 애널리틱스 요청 저장 (get_screen)
        self.captured_analytics_cards_request = None  # 캡처된 get_cards 요청 저장
        self.channel_id = None
        self.collected_data = []
        self.video_analytics_data = []  # 비디오 상세 정보 저장
        self.monitoring = False
        self.chrome_process = None
        self.pending_requests = {}  # 대기 중인 요청들 저장 (request_id -> request_data)
        self.collection_phase = "videos_list"  # "videos_list" → "video_analytics"
        
        # 다중 탭 애널리틱스 수집을 위한 설정
        self.analytics_tabs = [
            {
                'name': 'reach_viewers',
                'url_suffix': 'tab-reach_viewers/period-default',
                'api_endpoint': 'get_screen',
                'description': '노출수, CTR 데이터'
            },
            {
                'name': 'interest_viewers', 
                'url_suffix': 'tab-interest_viewers/period-default',
                'api_endpoint': 'get_cards',
                'description': '조회수, 시청시간, 구독자 증감 데이터'
            }
        ]
        self.current_tab_index = 0  # 현재 처리 중인 탭 인덱스
        self.collected_analytics_data = {}  # 탭별로 수집된 데이터 저장
        
        # 스레드 안전성을 위한 락
        import threading as thread_module
        self.tab_lock = thread_module.Lock()  # 탭 인덱스 변경 시 락
        
    def find_chrome_executable(self):
        """크롬 실행 파일 경로 찾기 (크로스 플랫폼)"""
        system = platform.system()
        
        if system == "Darwin":  # macOS
            possible_paths = [
                "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
                "/Applications/Google Chrome Canary.app/Contents/MacOS/Google Chrome Canary",
                "/Applications/Chromium.app/Contents/MacOS/Chromium",
                os.path.expanduser("~/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
            ]
        elif system == "Linux":
            possible_paths = [
                "/usr/bin/google-chrome",
                "/usr/bin/google-chrome-stable",
                "/usr/bin/chromium",
                "/usr/bin/chromium-browser",
                "/snap/bin/chromium",
                "/opt/google/chrome/chrome",
            ]
        elif system == "Windows":
            possible_paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe", 
                os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe"),
                r"C:\Users\%USERNAME%\AppData\Local\Google\Chrome\Application\chrome.exe"
            ]
        else:
            possible_paths = []
        
        # 경로에서 Chrome 찾기
        for path in possible_paths:
            expanded_path = os.path.expandvars(path)
            if os.path.exists(expanded_path):
                print(f"✅ Chrome 발견: {expanded_path}")
                return expanded_path
        
        # Windows에서는 레지스트리도 확인
        if system == "Windows":
            try:
                import winreg
                reg_path = r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\chrome.exe"
                with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, reg_path) as key:
                    chrome_path = winreg.QueryValue(key, "")
                    if os.path.exists(chrome_path):
                        print(f"✅ Chrome 발견 (레지스트리): {chrome_path}")
                        return chrome_path
            except Exception as e:
                print(f"레지스트리 검색 실패: {e}")
        
        # PATH에서 Chrome 찾기 (모든 플랫폼)
        chrome_names = []
        if system == "Darwin":
            # macOS에서는 PATH보다는 직접 경로를 우선시
            pass  
        elif system == "Linux":
            chrome_names = ["google-chrome", "google-chrome-stable", "chromium", "chromium-browser"]
        elif system == "Windows":
            chrome_names = ["chrome.exe", "chrome"]
        
        for name in chrome_names:
            chrome_path = shutil.which(name)
            if chrome_path:
                print(f"✅ Chrome 발견 (PATH): {chrome_path}")
                return chrome_path
        
        print(f"❌ Chrome을 찾을 수 없습니다 ({system})")
        return None
    
    def is_chrome_debug_running(self):
        """크롬이 디버그 모드로 실행 중인지 확인"""
        try:
            response = requests.get(f'http://localhost:{self.chrome_port}/json', timeout=2)
            return response.status_code == 200
        except:
            return False
    
    def kill_existing_chrome(self):
        """기존 크롬 프로세스 종료 (크로스 플랫폼)"""
        print("기존 크롬 프로세스를 종료하는 중...")
        system = platform.system()
        
        try:
            for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
                try:
                    proc_name = proc.info['name'].lower()
                    
                    # 플랫폼별 Chrome 프로세스 이름
                    chrome_names = []
                    if system == "Darwin":  # macOS
                        chrome_names = ['google chrome', 'chrome', 'chromium']
                    elif system == "Linux":
                        chrome_names = ['chrome', 'chromium', 'google-chrome', 'chromium-browser']
                    elif system == "Windows":
                        chrome_names = ['chrome.exe', 'chrome']
                    
                    # Chrome 프로세스인지 확인
                    is_chrome = any(name in proc_name for name in chrome_names)
                    
                    if is_chrome:
                        print(f"  Chrome 프로세스 종료: {proc.info['pid']} - {proc_name}")
                        proc.terminate()
                        
                except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                    pass
                    
        except Exception as e:
            print(f"프로세스 종료 중 오류: {e}")
        
        time.sleep(2)
    
    def get_program_profile_path(self):
        """프로그램 전용 Chrome 프로필 경로 반환 (크로스 플랫폼)"""
        system = platform.system()
        
        if system == "Darwin":  # macOS
            base_dir = os.path.expanduser("~/Library/Application Support")
        elif system == "Linux":
            base_dir = os.path.expanduser("~/.config")
        elif system == "Windows":
            base_dir = os.path.expanduser("~")
        else:
            base_dir = os.path.expanduser("~")
        
        profile_dir = os.path.join(base_dir, "YouTubeStudioMonitor", "chrome_profile")
        os.makedirs(profile_dir, exist_ok=True)
        print(f"Chrome 프로필 경로: {profile_dir}")
        return profile_dir
    
    def is_profile_logged_in(self, profile_path):
        """프로필이 Google 계정에 로그인되어 있는지 확인"""
        try:
            prefs_file = os.path.join(profile_path, "Default", "Preferences")
            if os.path.exists(prefs_file) and os.path.getsize(prefs_file) > 1000:
                return True
            return False
        except:
            return False
    
    def start_chrome_debug_mode(self):
        """크롬을 디버그 모드로 실행 (YouTube Studio로 바로 이동하지 않음)"""
        chrome_path = self.find_chrome_executable()
        if not chrome_path:
            print("크롬 실행 파일을 찾을 수 없습니다.")
            return False
        
        if not self.is_chrome_debug_running():
            self.kill_existing_chrome()
        
        user_data_dir = self.get_program_profile_path()
        is_logged_in = self.is_profile_logged_in(user_data_dir)
        
        if is_logged_in:
            print("기존 프로필을 사용합니다.")
        else:
            print("새 프로그램 전용 프로필을 생성합니다.")
        
        # YouTube Studio로 바로 이동하지 않고 Google 홈페이지로 시작
        chrome_args = [
            chrome_path,
            f"--remote-debugging-port={self.chrome_port}",
            f"--user-data-dir={user_data_dir}",
            f"--remote-allow-origins=http://localhost:{self.chrome_port}",
            "--no-first-run",
            "--no-default-browser-check",
            "--disable-web-security",
            "--disable-features=VizDisplayCompositor",
            "https://www.google.com"  # YouTube Studio 대신 Google 홈페이지로 시작
        ]
        
        try:
            print(f"크롬을 디버그 모드로 실행하는 중... (포트: {self.chrome_port})")
            self.chrome_process = subprocess.Popen(
                chrome_args,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=subprocess.CREATE_NEW_PROCESS_GROUP if os.name == 'nt' else 0
            )
            
            for i in range(10):
                if self.is_chrome_debug_running():
                    print("크롬이 성공적으로 시작되었습니다.")
                    return True
                time.sleep(1)
            
            print("크롬 시작 시간 초과")
            return False
            
        except Exception as e:
            print(f"크롬 실행 오류: {e}")
            return False

    def save_simple_excel(self, data, filename):
        """간단한 엑셀 파일 생성 (pandas 없이 openpyxl만 사용)"""
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'YouTube Data'
            
            if not data:
                print("❌ 저장할 데이터가 없습니다.")
                return
            
            # 헤더 생성
            headers = ['비디오 ID', '제목', '조회수', '게시일', '길이', '상태']
            
            # 헤더 작성
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num, value=header)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 데이터 작성
            for row_idx, item in enumerate(data, 2):
                worksheet.cell(row=row_idx, column=1, value=item.get('videoId', 'N/A'))
                worksheet.cell(row=row_idx, column=2, value=item.get('title', 'N/A'))
                worksheet.cell(row=row_idx, column=3, value=item.get('public_viewCount', 0))
                
                # 게시일 변환
                timestamp = item.get('timePublishedSeconds')
                if timestamp:
                    try:
                        dt = datetime.fromtimestamp(int(timestamp))
                        date_str = dt.strftime('%Y-%m-%d')
                    except:
                        date_str = 'N/A'
                else:
                    date_str = 'N/A'
                worksheet.cell(row=row_idx, column=4, value=date_str)
                
                # 길이 변환
                duration = item.get('lengthSeconds')
                if duration:
                    try:
                        seconds = int(duration)
                        minutes = seconds // 60
                        secs = seconds % 60
                        duration_str = f"{minutes}:{secs:02d}"
                    except:
                        duration_str = 'N/A'
                else:
                    duration_str = 'N/A'
                worksheet.cell(row=row_idx, column=5, value=duration_str)
                
                worksheet.cell(row=row_idx, column=6, value=item.get('privacy', 'N/A'))
            
            # 컬럼 너비 조정
            column_widths = [15, 50, 12, 12, 8, 12]
            for col_num, width in enumerate(column_widths, 1):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
            
            workbook.save(filename)
            print(f"✅ 간단한 엑셀 파일 생성: {filename}")
            
        except Exception as e:
            print(f"❌ 엑셀 저장 오류: {e}")

    # 나머지 메서드들은 기존과 동일하지만 pandas 사용 부분은 제거
    # 여기서는 핵심 메서드들만 포함하고 전체 코드는 별도로 생성

def main():
    print("YouTube Studio Monitor - 가벼운 버전")
    print("=" * 50)
    print("🎯 pandas 의존성을 제거한 안정적인 버전입니다!")
    
    monitor = YouTubeStudioMonitor()
    
    try:
        # 간단한 테스트
        chrome_path = monitor.find_chrome_executable()
        if chrome_path:
            print(f"✅ Chrome 경로 확인: {chrome_path}")
        else:
            print("❌ Chrome을 찾을 수 없습니다.")
            
        print("\n🚀 프로그램이 정상적으로 시작되었습니다!")
        print("실제 데이터 수집을 위해서는 Chrome 자동 실행 기능을 사용하세요.")
        
    except Exception as e:
        print(f"오류 발생: {e}")

if __name__ == "__main__":
    main() 
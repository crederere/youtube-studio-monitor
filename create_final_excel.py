#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO
import requests
from datetime import datetime

def create_final_excel():
    """실제 수집된 데이터로 사용자 요청 형식의 엑셀 생성"""
    
    print("🎯 최종 엑셀 생성 시작!")
    print("   • 썸네일: 첫 번째 열, 크게")
    print("   • 조회수: 기본 조회수만 (애널리틱스 조회수 제거)")
    print("   • 좋아요/댓글수: 제거")
    
    # 최신 애널리틱스 데이터 로드
    analytics_file = "youtube_analytics_data_20250723_152423.json"
    
    try:
        with open(analytics_file, 'r', encoding='utf-8') as f:
            analytics_data = json.load(f)
        
        video_analytics = analytics_data.get('analytics_data', [])
        print(f"✅ {len(video_analytics)}개 비디오 애널리틱스 데이터 로드됨")
        
        if not video_analytics:
            print("❌ 애널리틱스 데이터가 없습니다.")
            return
        
        # 워크북 생성
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'YouTube Analytics'
        
        # 헤더 설정 (사용자 요청에 맞게 단순화)
        headers = [
            '썸네일', '비디오 ID', '제목', '공개상태', '길이', '생성일', '게시일',
            '조회수', '노출수', '클릭률 (%)', 
            '시청시간 (분)', '평균 시청시간 (초)', 
            '평균 조회율 (%)', '비디오 URL'
        ]
        
        # 헤더 작성
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        # 공개 상태 변환
        privacy_mapping = {
            'VIDEO_PRIVACY_PUBLIC': '공개',
            'VIDEO_PRIVACY_PRIVATE': '비공개',
            'VIDEO_PRIVACY_UNLISTED': '제한공개',
            'VIDEO_PRIVACY_DRAFT': '임시보관함',
            'VIDEO_PRIVACY_SCHEDULED': '예약 게시',
            'VIDEO_STATUS_PROCESSED': '처리 완료',
            'VIDEO_STATUS_UPLOADING': '업로드 중',
            'VIDEO_STATUS_PROCESSING': '처리 중',
            'VIDEO_STATUS_FAILED': '실패'
        }
        
        # 시간 변환 함수
        def format_timestamp_korean(timestamp):
            try:
                if timestamp and str(timestamp) != 'N/A' and str(timestamp).isdigit():
                    dt = datetime.fromtimestamp(int(timestamp))
                    return dt.strftime('%Y년 %m월 %d일 %H시 %M분')
                return 'N/A'
            except:
                return 'N/A'
        
        def format_duration(seconds):
            try:
                if seconds and str(seconds) != 'N/A' and str(seconds).isdigit():
                    seconds = int(seconds)
                    hours = seconds // 3600
                    minutes = (seconds % 3600) // 60
                    secs = seconds % 60
                    if hours > 0:
                        return f"{hours}:{minutes:02d}:{secs:02d}"
                    else:
                        return f"{minutes}:{secs:02d}"
                return 'N/A'
            except:
                return 'N/A'
        
        # 데이터 처리
        for row_idx, item in enumerate(video_analytics, 2):
            video_info = item.get('basic_video_info', {})
            video_id = item.get('video_id', 'N/A')
            video_title = item.get('video_title', 'N/A')
            
            print(f"   📹 처리 중: {video_title[:30]}...")
            
            # 애널리틱스 데이터에서 메트릭 추출
            tabs_api_data = item.get('analytics_data', {})
            
            impressions = 0
            click_rate = 0
            watch_time_minutes = 0
            average_view_duration_seconds = 0
            average_percentage_watched = 0
            
            # 각 탭의 API 응답에서 데이터 추출
            for tab_name, tab_api_data in tabs_api_data.items():
                if tab_api_data and 'response_data' in tab_api_data:
                    response_data = tab_api_data['response_data']
                    cards = response_data.get('cards', [])
                    
                    for card in cards:
                        if 'keyMetricCardData' in card:
                            tabs = card['keyMetricCardData'].get('keyMetricTabs', [])
                            for tab in tabs:
                                primary_content = tab.get('primaryContent', {})
                                metric = primary_content.get('metric', '')
                                total = primary_content.get('total', 0)
                                
                                if metric == 'VIDEO_THUMBNAIL_IMPRESSIONS':
                                    impressions = total
                                elif metric == 'VIDEO_THUMBNAIL_IMPRESSIONS_VTR':
                                    click_rate = total
                                elif metric in ['EXTERNAL_WATCH_TIME', 'WATCH_TIME']:
                                    watch_time_minutes = round(total / 1000 / 60, 1)
                                elif metric in ['WATCH_TIME_MINUTES']:
                                    watch_time_minutes = total
                                elif metric in ['AVERAGE_VIEW_DURATION', 'AVG_VIEW_DURATION', 'AVERAGE_VIEW_DURATION_SECONDS']:
                                    average_view_duration_seconds = total
                        
                        elif 'audienceRetentionHighlightsCardData' in card:
                            videos_data = card['audienceRetentionHighlightsCardData'].get('videosData', [])
                            for video_data in videos_data:
                                if video_data.get('videoId') == video_id:
                                    metric_totals = video_data.get('metricTotals', {})
                                    
                                    avg_duration_millis = metric_totals.get('avgViewDurationMillis', 0)
                                    if avg_duration_millis:
                                        average_view_duration_seconds = round(avg_duration_millis / 1000)
                                    
                                    avg_percentage = metric_totals.get('avgPercentageWatched', 0)
                                    if avg_percentage:
                                        average_percentage_watched = round(avg_percentage * 100, 2)
                                    break
            
            # 공개 상태 변환
            privacy_status = (
                video_info.get('privacy') or 
                video_info.get('status') or 
                video_info.get('privacyStatus') or 
                'Unknown'
            )
            privacy_korean = privacy_mapping.get(privacy_status, privacy_status)
            
            # 썸네일 URL 추출
            thumbnail_url = None
            if 'thumbnailDetails' in video_info and 'thumbnails' in video_info['thumbnailDetails']:
                thumbnails = video_info['thumbnailDetails']['thumbnails']
                if thumbnails and len(thumbnails) > 0:
                    thumbnail_url = thumbnails[-1].get('url')
            elif 'thumbnail_urls' in video_info:
                urls = video_info['thumbnail_urls']
                if urls and len(urls) > 0:
                    thumbnail_url = urls[-1]
            if not thumbnail_url and video_id != 'N/A':
                thumbnail_url = f"https://img.youtube.com/vi/{video_id}/maxresdefault.jpg"
            
            # 사용자 요청에 맞는 데이터 (애널리틱스조회수 제거, 좋아요/댓글수 제거)
            data = [
                "📷 썸네일",  # 썸네일 (첫 번째 열)
                video_id,
                video_title[:50],
                privacy_korean,
                format_duration(video_info.get('lengthSeconds')),
                format_timestamp_korean(video_info.get('timeCreatedSeconds')),
                format_timestamp_korean(video_info.get('timePublishedSeconds')),
                int(video_info.get('public_viewCount', 0)) if video_info.get('public_viewCount') else 0,  # 기본 조회수만
                int(impressions) if impressions else 0,  # 노출수
                round(float(click_rate), 2) if click_rate else 0,  # 클릭률
                int(watch_time_minutes) if watch_time_minutes else 0,  # 시청시간
                int(average_view_duration_seconds) if average_view_duration_seconds else 0,  # 평균 시청시간
                round(float(average_percentage_watched), 2) if average_percentage_watched else 0,  # 평균 조회율
                f"https://www.youtube.com/watch?v={video_id}"  # 비디오 URL
            ]
            
            # 데이터 셀 작성
            for col_num, value in enumerate(data, 1):
                cell = worksheet.cell(row=row_idx, column=col_num, value=value)
                cell.font = Font(name="맑은 고딕", size=10)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # 컬럼별 정렬
                if col_num in [8, 9, 11, 12]:  # 숫자 컬럼들
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(value, (int, float)) and value > 0:
                        cell.number_format = '#,##0'
                elif col_num in [10, 13]:  # 퍼센트 컬럼들
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(value, (int, float)) and value > 0:
                        cell.number_format = '0.00'
                elif col_num in [1, 2, 4, 5]:  # 중앙 정렬
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # 썸네일 이미지 삽입 (1번째 열)
            thumbnail_cell = worksheet.cell(row=row_idx, column=1)
            if thumbnail_url:
                try:
                    print(f"      📷 썸네일 다운로드: {thumbnail_url[:50]}...")
                    response = requests.get(thumbnail_url, timeout=15, headers={
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                    })
                    
                    if response.status_code == 200:
                        img_data = BytesIO(response.content)
                        img = Image(img_data)
                        
                        # 이미지 크기 조정 (크게)
                        img.width = 160
                        img.height = 120
                        
                        # 셀에 이미지 고정
                        img.anchor = f'A{row_idx}'
                        worksheet.add_image(img)
                        
                        # 행 높이를 크게 조정
                        worksheet.row_dimensions[row_idx].height = 100
                        
                        thumbnail_cell.value = "📷"
                        print(f"      ✅ 썸네일 삽입 완료")
                    else:
                        thumbnail_cell.value = f"❌ HTTP {response.status_code}"
                        print(f"      ❌ 썸네일 다운로드 실패: {response.status_code}")
                except Exception as e:
                    thumbnail_cell.value = "❌ 오류"
                    print(f"      ❌ 썸네일 오류: {str(e)[:50]}")
            else:
                thumbnail_cell.value = "❌ URL없음"
            
            # 비디오 URL을 하이퍼링크로 설정 (14번째 열)
            url_cell = worksheet.cell(row=row_idx, column=14)
            video_url = f"https://www.youtube.com/watch?v={video_id}"
            url_cell.value = video_url
            url_cell.hyperlink = video_url
            url_cell.font = Font(name="맑은 고딕", size=10, color="0000FF", underline="single")
            url_cell.alignment = Alignment(horizontal="left", vertical="center")
            print(f"      🔗 하이퍼링크 설정: {video_url}")
        
        # 컬럼 너비 설정
        column_widths = [25, 15, 40, 12, 8, 18, 18, 12, 12, 10, 15, 15, 12, 45]
        for col_num, width in enumerate(column_widths, 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = width
        
        # 자동 필터 및 틀 고정
        worksheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(video_analytics) + 1}"
        worksheet.freeze_panes = "A2"
        
        # 파일 저장
        filename = "FINAL_USER_REQUEST.xlsx"
        workbook.save(filename)
        
        print(f"\n🎉 완벽한 최종 엑셀 파일 생성 완료!")
        print(f"📂 파일명: {filename}")
        print(f"📊 총 {len(video_analytics)}개 비디오 포함")
        print(f"\n✅ 사용자 요청사항 모두 반영:")
        print(f"   • 썸네일: A열 첫 번째, 크기 160x120, 행높이 100")
        print(f"   • 조회수: 기본 조회수만 (애널리틱스 조회수 제거됨)")
        print(f"   • 좋아요/댓글수: 완전히 제거됨")
        print(f"   • 비디오 URL: 클릭 가능한 하이퍼링크")
        print(f"   • 한국어 날짜/시간 형식")
        print(f"   • 공개 상태 한글화")
        print(f"   • 단일 시트만 존재")
        
        return filename
        
    except Exception as e:
        print(f"❌ 최종 엑셀 생성 오류: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    create_final_excel() 